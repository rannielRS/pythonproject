import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox
from decimal import *
from config import COLOR_CUERPO_PRINCIPAL, COLOR_BARRA_SUPERIOR,CONN_LOC,CURSOR_LOC,CONN_ZUN,CURSOR_ZUN
from PIL import Image, ImageTk
import openpyxl
import os
import subprocess
import tkinter.font as tkfont
from openpyxl.styles import Font, colors, fills, Alignment, PatternFill, NamedStyle
from copy import copy



class FormularioCalcImpDesign():
    
    def __init__(self, panel_principal):   
       
        #Definiendo variables
        # Variables de conexion
        self.connLoc = CONN_LOC
        self.cursorLoc = CURSOR_LOC
        self.connZun = CONN_ZUN
        self.cursorZun = CURSOR_ZUN

        if self.getPeriodo():
            
            # Definiendo controles de seleccion
            self.empSelec = ''
            self.tx_empleado_calceco = ttk.Entry(panel_principal, font=('Times', 14), width=10)
            self.tx_empleado_calceco.grid(row=0,column=0,padx=5,pady=5,ipadx=40)
            self.db_rm = []
            self.db_rm_str = StringVar()
            #Boton para buscar empleados        
            self.btn_bempleados_calceco = tk.Button(panel_principal, text="Buscar", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.actualizartreeCALCECO)
            self.btn_bempleados_calceco.place(x=250, y=2)
            #Para buscar por departamento
            #Label area
            self.tx_area_calceco = tk.Label(panel_principal, font=('Times', 14), width=20, bg=COLOR_CUERPO_PRINCIPAL, text='Departamento:')
            self.tx_area_calceco.place(x=350, y=5)

            #Combo departamento
            self.cb_area_calceco= ttk.Combobox(panel_principal, width=30)
            self.cb_area_calceco.bind('<<ComboboxSelected>>', self.actualizartreeCALCECO1)
            #self.cb_periodo.current(0)
            self.cb_area_calceco.place(x=520, y=5)

            style = ttk.Style()            
            style.configure('TLabelframe', background=COLOR_CUERPO_PRINCIPAL, borderwidth=2, bodercolor='black')
            style.configure('TLabelframe.Label', background=COLOR_CUERPO_PRINCIPAL)
            
            #Label frame para las invalidadnte
            self.lb_frame_calceco = ttk.Labelframe(panel_principal, text='Descuento por RM', style='TLabelframe')
            self.lb_frame_calceco.place(x=840, y=130, width=152,height=110)

            #Empleado label
            self.lb_sempleado_calceco = tk.Label(self.lb_frame_calceco, text='Emp.:', width=5, justify='center', bg=COLOR_CUERPO_PRINCIPAL, font=('Times', 11))
            self.lb_sempleado_calceco.grid(row=0,column=0, padx=5, pady=5)
            self.lb_sempleado_calceco.grid_propagate(False)            

            #Listado de empleados por id
            self.cb_emp_ninterno= ttk.Combobox(self.lb_frame_calceco, width=6)
            #self.cb_periodo.current(0)
            self.cb_emp_ninterno.grid(row=0,column=1, padx=5, pady=5)

            #LABEL COMBO empleado   
            self.lb_list_emp = tk.Label(self.lb_frame_calceco, text='Monto:', width=5, justify='center', bg=COLOR_CUERPO_PRINCIPAL, font=('Times', 11))
            self.lb_list_emp.grid(row=1,column=0, padx=5, pady=5)
            self.lb_list_emp.grid_propagate(False)

            #Descuento
            self.textoComentInv_calceco=ttk.Entry(self.lb_frame_calceco, font=('Times', 11), width=10, textvariable=self.db_rm_str)
            #self.textoComentInv_calceco.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
            self.textoComentInv_calceco.grid(row=1,column=1, padx=5, pady=5)
            self.textoComentInv_calceco.grid_propagate(False)

            #Definir imagenes de botones
            imagen_pil_btadd = Image.open("./imagenes/add.png")
            imagen_pil_btadd = imagen_pil_btadd.resize((20,20))
            imagen_btadd_tk = ImageTk.PhotoImage(imagen_pil_btadd)

            imagen_pil_btdel = Image.open("./imagenes/delete.png")
            imagen_pil_btdel = imagen_pil_btdel.resize((20,20))
            imagen_btdel_tk = ImageTk.PhotoImage(imagen_pil_btdel)

            imagen_pil_btlist = Image.open("./imagenes/list.png")
            imagen_pil_btlist = imagen_pil_btlist.resize((20,20))

            
            #Boton add inv        
            self.btn_addinv_calceco = tk.Button(self.lb_frame_calceco, text="\uf0c9",bd=0, image=imagen_btadd_tk, font=(
                'Times', 13), command=self.addInv)
            self.btn_addinv_calceco.image=imagen_btadd_tk
            self.btn_addinv_calceco.grid(row=2,column=0)

            #Boton elim inv        
            self.btn_elimInv_calceco = tk.Button(self.lb_frame_calceco, text="\uf0c9",bd=0, image=imagen_btdel_tk, font=(
                'Times', 13),  command=self.deleteInv)
            self.btn_elimInv_calceco.image=imagen_btdel_tk
            self.btn_elimInv_calceco.grid(row=2,column=1) 

            #Boton Informe/utilidades        
            self.btn_infUtil = tk.Button(panel_principal, text="Informe/utilidades", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.informeUtil)
            self.btn_infUtil.place(x=835, y=250)     

            #Boton Informe. x Depart.       
            self.btn_infUtilDep = tk.Button(panel_principal, text="Informe. x Depart.", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.informeDep)
            self.btn_infUtilDep.place(x=835, y=300)             


            #Boton Exportar dbf        
            self.btn_expdbf = tk.Button(panel_principal, text="Exportar dbf", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.expDBF)
            self.btn_expdbf.place(x=835, y=350)
            

            #Periodo del pago
            self.lx_PPlabel_mp = tk.Label(panel_principal, text="Mes de pago:", justify='right', bg=COLOR_CUERPO_PRINCIPAL, font=('Times', 12))
            self.lx_PPlabel_mp.place(x=850, y=80) 
            
            self.cb_periodo_calceco = ttk.Combobox(panel_principal, postcommand=self.getMesesA, width=12)
            #self.cb_periodo.current(0)
            self.cb_periodo_calceco.place(x=850, y=100) 
      
            

            #Treeview        
            columns = ('numero', 'ci', 'nombreap', 'devutil','segsoc', 'imping', 'descrm', 'neto')
            self.treeECalcEco = ttk.Treeview(panel_principal, height=16, columns=columns, show='headings')
            self.style = ttk.Style(self.treeECalcEco)
            self.style.configure('Treeview',rowheight=30)
            self.treeECalcEco.column('numero',width=80)
            self.treeECalcEco.column('ci',width=110)
            self.treeECalcEco.column('nombreap',width=200)            
            self.treeECalcEco.column('devutil',width=100)
            self.treeECalcEco.column('segsoc',width=80)
            self.treeECalcEco.column('imping',width=80)
            self.treeECalcEco.column('descrm',width=80)
            self.treeECalcEco.column('neto',width=80)

            self.treeECalcEco.heading(column='numero', text='No.')
            self.treeECalcEco.heading(column='ci', text='CI')
            self.treeECalcEco.heading(column='nombreap', text='Nombre y apellidos')            
            self.treeECalcEco.heading(column='devutil', text='Dev/Utili')
            self.treeECalcEco.heading(column='segsoc',text='Seg/Soc')
            self.treeECalcEco.heading(column='imping',text='Imp/Ing')
            self.treeECalcEco.heading(column='descrm',text='Desc/RM')
            self.treeECalcEco.heading(column='neto',text='Neto')            
            self.treeECalcEco.grid(row=1,column=0, columnspan=5,ipadx=5,padx=5,pady=5)
            self.treeECalcEco.bind('<<TreeviewSelect>>',self.selectEmp)
            self.actualizartreeCALCECO()  
            self.cargarEmpCB()  
            self.cargarDpto()   
        else:
            messagebox.showinfo('Notificación','Debe registrar un período de evaluación')

    def deleteInv(self):
        idemp = self.cb_emp_ninterno.get()
        if idemp:             
            if self.db_rm != []:
                for ecrm in self.db_rm:
                    finded = False
                    if ecrm[0]==idemp:
                        self.db_rm.remove(ecrm)
                        finded = True
                        messagebox.showinfo('Confirmación','Se eliminó el trabajador invalidado correctamente')
                        self.cb_emp_ninterno.set('')
                        self.db_rm_str.set('')
                if finded != True:
                    messagebox.showinfo('Sin acción','No existe ese trabajador en el registro')
            else:
                messagebox.showinfo('Error de validación','No existen elementos para eliminar')
        else:
            messagebox.showinfo('Información','Debe seleccionar el trabajador')
        


    def addInv(self):
        idemp = self.cb_emp_ninterno.get()
        if idemp:                 
            valor = self.textoComentInv_calceco.get()
            if valor != '':
                finded = False
                for ecrm in self.db_rm:
                    if ecrm[0] == idemp:
                        finded = True
                if finded:
                    return messagebox.showwarning('Registro repetido','Ese especialista ya se encentra invalidado')
                else:                    
                    self.db_rm.append((idemp,valor))
                    messagebox.showinfo('Confirmación','Se registró el trabajador invalidado correctamente')
                    self.cb_emp_ninterno.set('')
                    self.db_rm_str.set('')
            else:
                messagebox.showerror('Error de validación','Debe teclear el monto a descontar')

    
    def getDevengadoCalc(self, emp):
        query = "SELECT x.devengado FROM postgres.public.resumen_calculo_utilidades x where x.resumen_empleado_id ="+emp
        self.cursorLoc.execute(query)
        return self.cursorLoc.fetchone()[0]
    
    def getUtiliDist(self):
        queryUD="SELECT x.* FROM postgres.public.utilidades_distribucion x"
        self.cursorLoc.execute(queryUD)
        return self.cursorLoc.fetchone()


    #Cargar combo de periodo
    def cargarPeriodoOP(self): 
        slistp = self.getPeriodo() 
        options = []       
        for row in slistp:
            options.append(str(row[0])+"-"+str(row[1]))  

        self.cb_periodo_calceco['value']=options

    #Obtener resumen de registro de utilidades de un empleado
    def getResumenUtilEco(self,emp,utildev,salmespago):
        devengadot = utildev + salmespago
        #Seguridad social del salario del mes pago
        sal_mespago_5pcSC = 0
        sal_mespago_10pcSC = 0
        seg_social_mesSalario = 0
        # -5%
        if salmespago <= 15000:
            sal_mespago_5pcSC = salmespago * (5/100)
        else:
            sal_mespago_5pcSC = 15000 * (5/100)
        # -10%
        if salmespago <= 15000:
            sal_mespago_10pcSC = 0
        else:
            sal_mespago_10pcSC = (salmespago-15000) * (10/100)
        seg_social_mesSalario = sal_mespago_5pcSC + sal_mespago_10pcSC
        # Seguridad social total
        sst_5pc = 0
        sst_10pc = 0
        sst = 0
        sst_diferencia = 0
        # -5%
        if devengadot <= 15000:
            sst_5pc = devengadot * (5/100)
        else:
            sst_5pc = 15000 * (5/100)
        # -10%
        if devengadot <= 15000:
            sst_10pc = 0
        else:
            sst_10pc = (devengadot-15000) * (10/100)
        sst = sst_5pc + sst_10pc
        sst_diferencia = sst - seg_social_mesSalario
        #Impuestos sobre ingresos del mes pago
        iimp_3pc = 0
        iimp_5pc = 0
        iimp_7_5pc = 0
        iimp_10pc = 0
        iimp_15pc = 0
        iimp_20pc = 0
        iit_mespago = 0
        # -3%
        if salmespago > 3260:
            if salmespago>9510:
                iimp_3pc=(9510-3260)*(3/100)
            else:
                iimp_3pc=(salmespago-3260)*(3/100)
        else:
            iimp_3pc=0
        # -5%
        if salmespago > 9510:
            if salmespago>15000:
                iimp_5pc=(15000-9510)*(5/100)
            else:
                iimp_5pc=(salmespago-9510)*(5/100)
        else:
            iimp_5pc=0
        # -7.5%
        if salmespago > 15000:
            if salmespago>20000:
                iimp_7_5pc=(20000-15000)*(7.5/100)
            else:
                iimp_7_5pc=(salmespago-15000)*(7.5/100)
        else:
            iimp_7_5pc=0
        # -10%
        if salmespago > 20000:
            if salmespago>25000:
                iimp_10pc=(25000-20000)*(10/100)
            else:
                iimp_10pc=(salmespago-20000)*(10/100)
        else:
            iimp_10pc=0
        # -15%
        if salmespago > 25000:
            if salmespago>30000:
                iimp_15pc=(30000-25000)*(15/100)
            else:
                iimp_15pc=(salmespago-25000)*(15/100)
        else:
            iimp_15pc=0
        # -20%
        if salmespago > 30000:
            iimp_20pc=(salmespago-30000)*(20/100)
        else:
            iimp_20pc=0
        iit_mespago = iimp_3pc+iimp_5pc+iimp_7_5pc+iimp_10pc+iimp_15pc+iimp_20pc
        #Impuestos sobre ingresos totales
        iit_3pc = 0
        iit_5pc = 0
        iit_7_5pc = 0
        iit_10pc = 0
        iit_15pc = 0
        iit_20pc = 0
        iit = 0
        iit_diferencia = 0

        # -3%
        if devengadot > 3260:
            if devengadot>9510:
                iit_3pc=(9510-3260)*(3/100)
            else:
                iit_3pc=(devengadot-3260)*(3/100)
        else:
            iit_3pc=0
        # -5%
        if devengadot > 9510:
            if devengadot>15000:
                iit_5pc=(15000-9510)*(5/100)
            else:
                iit_5pc=(devengadot-9510)*(5/100)
        else:
            iit_5pc=0
        # -7.5%
        if devengadot > 15000:
            if devengadot>20000:
                iit_7_5pc=(20000-15000)*(7.5/100)
            else:
                iit_7_5pc=(devengadot-15000)*(7.5/100)
        else:
            iit_7_5pc=0
        # -10%
        if devengadot > 20000:
            if devengadot>25000:
                iit_10pc=(25000-20000)*(10/100)
            else:
                iit_10pc=(devengadot-20000)*(10/100)
        else:
            iit_10pc=0
        # -15%
        if devengadot > 25000:
            if devengadot>30000:
                iit_15pc=(30000-25000)*(15/100)
            else:
                iit_15pc=(devengadot-25000)*(15/100)
        else:
            iit_15pc=0
        # -20%
        if devengadot > 30000:
            iit_20pc=(devengadot-30000)*(20/100)
        else:
            iit_20pc=0
        iit = iit_3pc+iit_5pc+iit_7_5pc+iit_10pc+iit_15pc+iit_20pc
        iit_diferencia = iit-iit_mespago

        #Neto a cobrar
        respMat = 0
        tupla = []
        for rm in self.db_rm:
            idemp= emp
            if idemp in rm:
                tupla=rm
        if tupla:
            respMat = float(tupla[1])
        neto_salario = utildev-sst_diferencia-iit_diferencia-respMat
        return [utildev,sst_diferencia,iit_diferencia,respMat,neto_salario]

    
    #Definiendo tree view de periodo    
    def informeUtil(self):   
        path = "file/utilidades_disteco.xlsx"
        row = 6 
        controw = 1
        self.limpiarExcel(row,path)   
        wb = openpyxl.load_workbook(path)
        sheet = wb.active

        if self.cb_periodo_calceco.get() == '':
            return messagebox.showwarning('Validación de información','Debe seleccionar el mes del pago')
        
        self.limpiarHistUtil()
        montoDistribuir = Decimal(self.getUtiliDist()[2])
        sheet['I3'] = montoDistribuir
        sheet['I3'].number_format = '#,##0.00'
        
        alignmentText = Alignment(horizontal=LEFT)
        alignmentNumber = Alignment(horizontal=CENTER)
        text_format = Font(
        bold = False,
        name = 'Calibri',
        size = '0',
        color = colors.BLACK )   
        number_format = Font(
        bold = False,
        name = 'Calibri',
        size = '0',
        color = colors.BLACK) 
        sheet['N5'] =  self.getPeriodo()[0][1]
        sheet['O5'] =  self.getPeriodo()[1][1]
        sheet['P5'] =  self.getPeriodo()[2][1]

        sheet['G4'] =  self.getPeriodo()[0][1]
        sheet['I4'] =  self.getPeriodo()[1][1]
        sheet['K4'] =  self.getPeriodo()[2][1]

        queryP="SELECT a.area,emp.id,emp.ci,emp.nombreap,emp.escalas,emp.thoraria  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id ORDER BY a.id"
        self.cursorLoc.execute(queryP)
        listEmp = self.cursorLoc.fetchall()
        for empleado in listEmp:
            

            sheet['B'+str(row)] =  controw
            sheet['B'+str(row)].font +=  number_format
            sheet['B'+str(row)].alignment += alignmentNumber
            sheet['C'+str(row)] =  empleado[1]
            sheet['C'+str(row)].font +=  text_format
            sheet['C'+str(row)].alignment += alignmentText
            sheet['D'+str(row)] =  empleado[2]
            sheet['D'+str(row)].font +=  text_format
            sheet['D'+str(row)].alignment += alignmentText
            sheet['E'+str(row)] =  empleado[3]
            sheet['E'+str(row)].font +=  text_format
            sheet['E'+str(row)].alignment += alignmentText
            sheet['F'+str(row)] =  empleado[4]
            sheet['F'+str(row)].font +=  text_format
            sheet['F'+str(row)].alignment += alignmentText

                
            sheet['G'+str(row)] =  '0'
            sheet['G'+str(row)].font +=  number_format
            sheet['G'+str(row)].alignment += alignmentNumber                
            sheet['H'+str(row)] =  '0'
            sheet['H'+str(row)].font +=  number_format
            sheet['H'+str(row)].alignment += alignmentNumber
            sheet['I'+str(row)] =  '0'
            sheet['I'+str(row)].font +=  number_format
            sheet['I'+str(row)].alignment += alignmentNumber
            sheet['J'+str(row)] =  '0'
            sheet['J'+str(row)].font +=  number_format
            sheet['J'+str(row)].alignment += alignmentNumber
            sheet['K'+str(row)] =  '0'
            sheet['K'+str(row)].font +=  number_format
            sheet['K'+str(row)].alignment += alignmentNumber
            sheet['L'+str(row)] =  'NE'
            sheet['L'+str(row)].font +=  number_format
            sheet['L'+str(row)].alignment += alignmentNumber
            sheet['M'+str(row)] =  'NE'
            sheet['M'+str(row)].font +=  number_format
            sheet['M'+str(row)].alignment += alignmentNumber
            sheet['N'+str(row)] =  'NE'
            sheet['N'+str(row)].font +=  number_format
            sheet['N'+str(row)].alignment += alignmentNumber
            sheet['N'+str(row)].number_format = '#,#0.0'
            sheet['O'+str(row)] =  '0'
            sheet['O'+str(row)].font +=  number_format
            sheet['O'+str(row)].alignment += alignmentNumber
            sheet['O'+str(row)].number_format = '#,#0.0'
            sheet['P'+str(row)] =  '0'
            sheet['P'+str(row)].font +=  number_format
            sheet['P'+str(row)].alignment += alignmentNumber
            sheet['P'+str(row)].number_format = '#,#0.0'
            sheet['Q'+str(row)] =  '0'
            sheet['Q'+str(row)].font +=  number_format
            sheet['Q'+str(row)].alignment += alignmentNumber
            sheet['Q'+str(row)].number_format = '#,##0.00'
            sheet['R'+str(row)] =  '0'
            sheet['R'+str(row)].font +=  number_format
            sheet['R'+str(row)].alignment += alignmentNumber
            sheet['R'+str(row)].number_format = '#,##0.00'
            sheet['S'+str(row)] =  '0'
            sheet['S'+str(row)].font +=  number_format
            sheet['S'+str(row)].alignment += alignmentNumber
            sheet['S'+str(row)].number_format = '#,##0.00'
            sheet['T'+str(row)] =  '0'
            sheet['T'+str(row)].font +=  number_format
            sheet['T'+str(row)].alignment += alignmentNumber
            sheet['T'+str(row)].number_format = '#,##0.00'
            sheet['U'+str(row)] =  '0'
            sheet['U'+str(row)].font +=  number_format
            sheet['U'+str(row)].alignment += alignmentNumber
            sheet['U'+str(row)].number_format = '#,##0.00'

            #Informacion de impuestos
            sheet['X'+str(row)].font +=  number_format
            sheet['X'+str(row)].alignment += alignmentNumber
            sheet['X'+str(row)].number_format = '#,##0.00'
            sheet['Y'+str(row)].font +=  number_format
            sheet['Y'+str(row)].alignment += alignmentNumber
            sheet['Y'+str(row)].number_format = '#,##0.00'
            sheet['Z'+str(row)].font +=  number_format
            sheet['Z'+str(row)].alignment += alignmentNumber
            sheet['Z'+str(row)].number_format = '#,##0.00'
            sheet['AA'+str(row)].font +=  number_format
            sheet['AA'+str(row)].alignment += alignmentNumber
            sheet['AA'+str(row)].number_format = '#,##0.00'
            sheet['AA'+str(row)]=f'=IF(X{row}<=15000,X{row}*AA$5,15000*AA$5)'
            sheet['AB'+str(row)].font +=  number_format
            sheet['AB'+str(row)].alignment += alignmentNumber
            sheet['AB'+str(row)].number_format = '#,##0.00'
            sheet['AB'+str(row)]=f'=IF(X{row}<=15000,0,(X{row}-15000)*AB$5)'
            sheet['AC'+str(row)].font +=  number_format
            sheet['AC'+str(row)].alignment += alignmentNumber
            sheet['AC'+str(row)].number_format = '#,##0.00'
            sheet['AC'+str(row)]=f'=SUM(AA{row}:AB{row})'
            sheet['AD'+str(row)].font +=  number_format
            sheet['AD'+str(row)].alignment += alignmentNumber
            sheet['AD'+str(row)].number_format = '#,##0.00'
            sheet['AD'+str(row)]=f'=IF(Z{row}<=15000,Z{row}*AD$5,15000*AD$5)'
            sheet['AE'+str(row)].font +=  number_format
            sheet['AE'+str(row)].alignment += alignmentNumber
            sheet['AE'+str(row)].number_format = '#,##0.00'
            sheet['AE'+str(row)]=f'=IF(Z{row}<=15000,0,(Z{row}-15000)*AE$5)'
            sheet['AF'+str(row)].font +=  number_format
            sheet['AF'+str(row)].alignment += alignmentNumber
            sheet['AF'+str(row)].number_format = '#,##0.00'
            sheet['AF'+str(row)]=f'=SUM(AD{row}:AE{row})'
            sheet['AG'+str(row)].font +=  number_format
            sheet['AG'+str(row)].alignment += alignmentNumber
            sheet['AG'+str(row)].number_format = '#,##0.00'
            sheet['AG'+str(row)]=f'=SUM(AF{row}-AC{row})'
            sheet['AH'+str(row)].font +=  number_format
            sheet['AH'+str(row)].alignment += alignmentNumber
            sheet['AH'+str(row)].number_format = '#,##0.00'
            sheet['AH'+str(row)]=f'=IF(X{row}>3260,IF(X{row}>9510,(9510-3260)*AH$5,(X{row}-3260)*AH$5),0)'
            sheet['AI'+str(row)].font +=  number_format
            sheet['AI'+str(row)].alignment += alignmentNumber
            sheet['AI'+str(row)].number_format = '#,##0.00'
            sheet['AI'+str(row)]=f'=IF(X{row}>9510,IF(X{row}>15000,(15000-9510)*AI$5,(X{row}-9510)*AI$5),0)'
            sheet['AJ'+str(row)].font +=  number_format
            sheet['AJ'+str(row)].alignment += alignmentNumber
            sheet['AJ'+str(row)].number_format = '#,##0.00'
            sheet['AJ'+str(row)]=f'=IF(X{row}>15000,IF(X{row}>20000,(20000-15000)*AJ$5,(X{row}-15000)*AJ$5),0)'
            sheet['AK'+str(row)].font +=  number_format
            sheet['AK'+str(row)].alignment += alignmentNumber
            sheet['AK'+str(row)].number_format = '#,##0.00'
            sheet['AK'+str(row)]=f'=IF(X{row}>20000,IF(X{row}>25000,(25000-20000)*AK$5,(X{row}-20000)*AK$5),0)'
            sheet['AL'+str(row)].font +=  number_format
            sheet['AL'+str(row)].alignment += alignmentNumber
            sheet['AL'+str(row)].number_format = '#,##0.00'
            sheet['AL'+str(row)]=f'=IF(X{row}>25000,IF(X{row}>30000,(30000-25000)*AL$5,(X{row}-25000)*AL$5),0)'
            sheet['AM'+str(row)].font +=  number_format
            sheet['AM'+str(row)].alignment += alignmentNumber
            sheet['AM'+str(row)].number_format = '#,##0.00'
            sheet['AM'+str(row)]=f'=IF(X{row}>30000,(X{row}-30000)*AM$5,0)'
            sheet['AN'+str(row)].font +=  number_format
            sheet['AN'+str(row)].alignment += alignmentNumber
            sheet['AN'+str(row)].number_format = '#,##0.00'
            sheet['AN'+str(row)]=f'=SUM(AH{row}:AM{row})'
            sheet['AO'+str(row)].font +=  number_format
            sheet['AO'+str(row)].alignment += alignmentNumber
            sheet['AO'+str(row)].number_format = '#,##0.00'
            sheet['AO'+str(row)]=f'=IF(Z{row}>3260,IF(Z{row}>9510,(9510-3260)*AO$5,(Z{row}-3260)*AO$5),0)'
            sheet['AP'+str(row)].font +=  number_format
            sheet['AP'+str(row)].alignment += alignmentNumber
            sheet['AP'+str(row)].number_format = '#,##0.00'
            sheet['AP'+str(row)]=f'=IF(Z{row}>9510,IF(Z{row}>15000,(15000-9510)*AP$5,(Z{row}-9510)*AP$5),0)'
            sheet['AQ'+str(row)].font +=  number_format
            sheet['AQ'+str(row)].alignment += alignmentNumber
            sheet['AQ'+str(row)].number_format = '#,##0.00'
            sheet['AQ'+str(row)]=f'=IF(Z{row}>15000,IF(Z{row}>20000,(20000-15000)*AQ$5,(Z{row}-15000)*AQ$5),0)'
            sheet['AR'+str(row)].font +=  number_format
            sheet['AR'+str(row)].alignment += alignmentNumber
            sheet['AR'+str(row)].number_format = '#,##0.00'
            sheet['AR'+str(row)]=f'=IF(Z{row}>20000,IF(Z{row}>25000,(25000-20000)*AR$5,(Z{row}-20000)*AR$5),0)'
            sheet['AS'+str(row)].font +=  number_format
            sheet['AS'+str(row)].alignment += alignmentNumber
            sheet['AS'+str(row)].number_format = '#,##0.00'
            sheet['AS'+str(row)]=f'=IF(Z{row}>25000,IF(Z{row}>30000,(30000-25000)*AS$5,(Z{row}-25000)*AS$5),0)'
            sheet['AT'+str(row)].font +=  number_format
            sheet['AT'+str(row)].alignment += alignmentNumber
            sheet['AT'+str(row)].number_format = '#,##0.00'
            sheet['AT'+str(row)]=f'=IF(Z{row}>30000,(Z{row}-30000)*AT$5,0)'
            sheet['AU'+str(row)].font +=  number_format
            sheet['AU'+str(row)].alignment += alignmentNumber
            sheet['AU'+str(row)].number_format = '#,##0.00'
            sheet['AU'+str(row)]=f'=SUM(AO{row}:AT{row})'
            sheet['AV'+str(row)].font +=  number_format
            sheet['AV'+str(row)].alignment += alignmentNumber
            sheet['AV'+str(row)].number_format = '#,##0.00'
            sheet['AV'+str(row)]=f'=AU{row}-AN{row}'
            sheet['AW'+str(row)].font +=  number_format
            sheet['AW'+str(row)].alignment += alignmentNumber
            sheet['AW'+str(row)].number_format = '#,##0.00'
            sheet['AX'+str(row)].font +=  number_format
            sheet['AX'+str(row)].alignment += alignmentNumber
            sheet['AX'+str(row)].number_format = '#,##0.00'
            sheet['AX'+str(row)]=f'=Y{row}-AG{row}-AV{row}-AW{row}'

            if self.getPagosTM(empleado[1])['pago_tm_mn'] == 1:
                sheet['AY'+str(row)] = 'TM'
            else:
                sheet['AY'+str(row)] = ''

            idsperiodos =  []
            periodos = list(self.getPeriodo())
            idsperiodos.append(periodos[0][0])
            idsperiodos.append(periodos[1][0])
            idsperiodos.append(periodos[2][0])  
            vacacionesmAnt1 = 0
            vacacionesm1 = 0
            vacacionesm2 = 0
            vacacionesm3 = 0
            vacaciones = self.getVacacionesMT(empleado[1])                   
            for v in vacaciones:               
                if v[2] in idsperiodos:
                    if idsperiodos.index(v[2]) == 0:
                        vacacionesm1 =  v[0] 
                    if idsperiodos.index(v[2]) == 1:
                        vacacionesm2 =  v[0]
                    if idsperiodos.index(v[2]) == 2:
                        vacacionesm3 =  v[0]
                else:
                    vacacionesmAnt1 =  v[0]

            salariomes1 = 0
            horasmes1 = 0
            salariomes2 = 0
            horasmes2 = 0
            salariomes3 = 0 
            horasmes3 = 0           
            salarios = list(self.getpagoSalMT(empleado[1]))
            for sal in salarios:
                if idsperiodos.index(sal[3]) == 0:
                    salariomes1 =  sal[0] - sal[2]
                    horasmes1 = sal[1]
                if idsperiodos.index(sal[3]) == 1:
                    salariomes2 =  sal[0] - sal[2]
                    horasmes2 = sal[1]
                if idsperiodos.index(sal[3]) == 2:
                    salariomes3 =  sal[0] - sal[2]
                    horasmes3 = sal[1]

            sheet['G'+str(row)] = horasmes1
            sheet['I'+str(row)] = horasmes2
            sheet['K'+str(row)] = horasmes3

            sheet['H'+str(row)] = Decimal(vacacionesmAnt1)+Decimal(vacacionesm1)+Decimal(salariomes1)
            sheet['J'+str(row)] = Decimal(vacacionesm2)+Decimal(salariomes2)
            sheet['L'+str(row)] = Decimal(vacacionesm3)+Decimal(salariomes3)

            sheet['M'+str(row)] = round((((Decimal(vacacionesmAnt1)+Decimal(vacacionesm1)+Decimal(salariomes1))+(Decimal(vacacionesm2)+Decimal(salariomes2))+(Decimal(vacacionesm3)+Decimal(salariomes3)))/3),2)
            
            mes1 = round((self.obtenerEvaCond(empleado[1],periodos[0][0])[0]),1)
            mes2 = round((self.obtenerEvaCond(empleado[1],periodos[1][0])[0]),1)
            mes3 = round((self.obtenerEvaCond(empleado[1],periodos[2][0])[0]),1)

            # if salariomes1 == 0 and mes1 != 'NE':
            #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} no tiene pago en el mes {periodos[0][1]} debe estar como NE')
            # if salariomes1 != 0 and mes1 == 'NE':
            #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} debe ser evaluado en el mes {periodos[0][1]} debido a que cuenta con un salario devengado')
            sheet['N'+str(row)] = mes1

            # if salariomes2 == 0 and mes2 != 'NE':
            #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} no tiene pago en el mes {periodos[1][1]} debe estar como NE')
            # if salariomes2 != 0 and mes2 == 'NE':
            #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} debe ser evaluado en el mes {periodos[1][1]} debido a que cuenta con un salario devengado')              
            sheet['O'+str(row)] = mes2
            
            # if salariomes3 == 0 and mes3 != 'NE':
            #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} no tiene pago en el mes {periodos[2][1]} debe estar como NE')
            # if salariomes3 != 0 and mes3 == 'NE':
            #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} debe ser evaluado en el mes {periodos[2][1]} debido a que cuenta con un salario devengado')
            sheet['P'+str(row)] = mes3
            


            promEva = self.calcCoeficienteEva(empleado[1])
            sheet['Q'+str(row)] = promEva
            if promEva <= 2:
                sheet['R'+str(row)] = 0
            else:
                sheet['R'+str(row)] = promEva
            
            #Calculo del salario base de cada trabajador
            sheet['S'+str(row)] = f'=M{row}*R{row}' 
                   

            row += 1
            controw += 1

        sheet['K3'] = f'=SUM(S6:S{row+6})'
        sheet['M3'] = f'=I3/K3'
        sheet['M3'].number_format ='#,##0.00000'

        for i in range(6,row):
            sheet['T'+str(i)] = sheet['M3'].value
            sheet['U'+str(i)] = self.getDevengadoCalc("'"+sheet['C'+str(i)].value+"'")#f'=rounddown((S{i}*T{i}),2)'
            #iMPUESTOS
            periodoPago = self.cb_periodo_calceco.get().split('-')[0]
            salarioMesp = self.getSalarioNomMes("'"+sheet['C'+str(i)].value+"'",periodoPago)
            if salarioMesp is not None:
                sheet['X'+str(i)] = salarioMesp['deveng_salario']
            else:
                sheet['X'+str(i)] = 0
            sheet['Y'+str(i)] = f'=U{i}'  
            sheet['Z'+str(i)] = f'=X{i}+Y{i}'
            #[utildev,sst_diferencia,iit_diferencia,respMat,neto_salario]
            resumencalc_eco = self.getResumenUtilEco(sheet['C'+str(i)].value,float(self.getDevengadoCalc("'"+sheet['C'+str(i)].value+"'")),float(sheet['X'+str(i)].value))
            queryinsert_hutil = "INSERT INTO postgres.public.utilidades_printhist(codigo_empleado,name_utilidad,ci,nombap,devengado_util,aporte_ss,imp_ingp,descuento_rm,neto_cobrar,area) \
                VALUES('"+sheet['C'+str(i)].value+"','"+self.getUtiliDist()[1]+"_"+str(self.getUtiliDist()[3])+"','"+sheet['D'+str(i)].value+"','"+sheet['E'+str(i)].value+"',"+str(round(resumencalc_eco[0],2))+","+str(round(resumencalc_eco[1],2))+","+str(round(resumencalc_eco[2],2))+","+str(round(resumencalc_eco[3],2))+","+str(round(resumencalc_eco[4],2))+",'"+self.getDepartamentoEmp(sheet['C'+str(i)].value)[0]+"')"
            self.cursorLoc.execute(queryinsert_hutil)
            self.connLoc.commit()        
        wb.save(path)
        self.actualizartreeCALCECO()
        separador = os.path.sep
        dir_actual = os.path.dirname(os.path.abspath(__file__))
        dir = separador.join(dir_actual.split(separador)[:-1])
        dirfile = separador.join(path.split(separador))
        command =  ['open', dir+separador+dirfile]
        subprocess.run(command,shell=False)

        #self.convert_xlsx_to_pdf(path,"utilidades_disteco")
        
        
    def convert_xlsx_to_pdf(self,xlsx_file,nombreA=''):
        try:
            subprocess.run(["libreoffice24.2", "--headless", "--convert-to", "pdf", xlsx_file])
            separador = os.path.sep
            dir_actual = os.path.dirname(os.path.abspath(__file__))
            dir = separador.join(dir_actual.split(separador)[:-1])
            #dirfile = separador.join(xlsx_file.split(separador))
            command =  ['open', dir+separador+nombreA+'.pdf']
            subprocess.run(command,shell=False)
            print("Done!")

        except Exception as e:
            print("Error:", e)

    #  ----------------------------------------
    #  Funcion de Ayuda
    #  Mantener estilos y formatos
    #  Entre celdas
    #  requiere importar copy
    #  ----------------------------------------
    def copyStyle(newCell, cell):
        if cell.has_style:
            newCell.style = copy(cell.style)
            newCell.font = copy(cell.font)
            newCell.border = copy(cell.border)
            newCell.fill = copy(cell.fill)
            newCell.number_format = copy(cell.number_format)
            newCell.protection = copy(cell.protection)
            newCell.alignment = copy(cell.alignment)


    #  ----------------------------------------
    #  Funcion de Ayuda
    #  Para copiar valores entre columnas
    #  Con numero de filas variables 
    #  ----------------------------------------
    def copyFromTo(Fro, To, start=1, num=1, sheet=''):

        i = start                               # indica la fila incial                         

        while (True):                           # iterar
            a = sheet[Fro + str(i)].value          # se coge el valor de la celda origen
            if a == None:                       # cuando no hay nada en la celda
                return                          # se sale de while con return
            else:
                # Esta parte para este caso indica
                # 1 no tiene sentido dividir pero
                # al mismo tiempo se le puede emplear
                # cuando se copian valores del
                # tipo string o date
                if num != 1:                    
                    sheet[To + str(i)].value = a / num
                else:
                    sheet[To + str(i)].value = a

                # se mantienen los estilos y formatos
                copyStyle(sheet[To + str(i)], sheet[Fro + str(i)])

                # Efectos de control
                print(i,':',a, sheet[To + str(i)].value)

            # Se actualiza para la siguiente fila
            i += 1

    def selectEmp(self,event):
        selectItem=self.treeECalcEco.item(self.treeECalcEco.selection())
        idemp=str(selectItem['values'][0])
        self.cb_emp_ninterno.set(idemp.replace("'",""))
        self.db_rm_str.set(selectItem['values'][6])

    def limpiarHistUtil(self):
        queryDemp = "DELETE FROM postgres.public.utilidades_printhist WHERE name_utilidad='"+self.getUtiliDist()[1]+"_"+str(self.getUtiliDist()[3])+"'"
        self.cursorLoc.execute(queryDemp)
        self.connLoc.commit()
    
    def getPagosTM(self,emp):
        querysalnom = "SELECT ppt.pago_tm_mn, ppt.cuenta_tm_cup FROM ZUNpr.dbo.p_pagos_tm AS ppt WHERE ppt.no_interno="+str(emp)
        self.cursorZun.execute(querysalnom)
        result = self.cursorZun.fetchone()
        return result

    def getSalarioNomMes(self,emp,mes):
        querysalnom = "SELECT ns.deveng_salario FROM ZUNpr.dbo.nomina_sal AS ns WHERE no_interno="+str(emp)+" AND id_periodo="+str(mes)
        self.cursorZun.execute(querysalnom)
        result = self.cursorZun.fetchone()
        return result
    
    def informeDep(self):
        pass
    
    def getDepartamentoEmp(self,emp):
        query = "SELECT a.area FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id WHERE emp.id='"+emp+"' ORDER BY a.id"
        self.cursorLoc.execute(query)
        return self.cursorLoc.fetchone()
        

    def getVacacionesMT(self, empleado):
        mtvacaciones = []
        querygetVaca="SELECT v.importe_total,v.tiempo_tota,p.id  FROM postgres.public.periodo p \
        INNER JOIN postgres.public.vacacionesp AS v ON p.id = v.vacacionesp_periodo_id where v.vacacionesp_empleado_id  = '"+empleado+"' order by p.id asc" 
        self.cursorLoc.execute(querygetVaca)
        vacacionesList = self.cursorLoc.fetchall()
        for vac in vacacionesList:
            if self.getDestajo(empleado)[1]:
                tiempo=float(vac[1])
                if self.getTarifaZun(empleado[0],vac[2]) is not None:
                    tarifaH = self.getTarifaZun(empleado[0],vac[2])[0]
                else:
                    tarifaH=float(self.getDestajo(empleado)[0])
                calcVaca = tiempo*tarifaH*8
                mtvacaciones.append((round(calcVaca,2), vac[1],vac[2]))
            else:
                mtvacaciones.append((vac[0], vac[1],vac[2]))
        return mtvacaciones

    def getDestajo(self,empleado):
        querygetSal="SELECT e.thoraria,e.destajo FROM postgres.public.empleado AS e \
            where e.id  = '"+empleado+"'"
        self.cursorLoc.execute(querygetSal)
        result = self.cursorLoc.fetchone()
        return result

    def getTarifaZun(self,empleado,periodo):
        querygetTarifa="SELECT x.tarifa_horaria FROM ZUNpr.dbo.h_empleado x \
            where x.no_interno  = '"+empleado+"' AND x.id_peri = "+str(periodo)
        self.cursorZun.execute(querygetTarifa)
        result = self.cursorZun.fetchone()
        return result

    def obtenerEvaCond(self,emp,periodo):
        queryP="SELECT te.peso FROM postgres.public.evaluacion AS e  INNER JOIN postgres.public.tipo_evaluacion AS te ON e.evaluacion_tipoevaluacion_id = te.id where e.evaluacion_empleado_id='"+str(emp)+"' and e.evaluacion_perio_id='"+str(periodo)+"'"
        #print(queryP)
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchone()
        
    def actualizartreeCALCECO(self):
        self.treeECalcEco.delete(*self.treeECalcEco.get_children())         
        queryEmpL=''
        if self.tx_empleado_calceco.get() != '' and self.cb_area_calceco.get() == '':
            queryEmpL="SELECT x.* FROM postgres.public.utilidades_printhist x where x.nombap like '%"+self.tx_empleado_calceco.get().upper()+"%'"
        elif self.cb_area_calceco.get() != '' and self.tx_empleado_calceco.get() == '':
            queryEmpL="SELECT x.* FROM postgres.public.utilidades_printhist x where x.area = '"+self.cb_area_calceco.get()+"'"
        elif self.tx_empleado_calceco.get() != '' and self.cb_area_calceco.get() != '':
            queryEmpL="SELECT x.* FROM postgres.public.utilidades_printhist x where x.nombap like '%"+self.tx_empleado_calceco.get().upper()+"%' and x.area = '"+self.cb_area_calceco.get()+"'"
        else:
            queryEmpL='SELECT x.* FROM postgres.public.utilidades_printhist x'        
         
        self.cursorLoc.execute(queryEmpL)

        slistEmp = self.cursorLoc.fetchall()            
        for emp in slistEmp:
            if emp[8] != 0:
                self.db_rm.append((emp[1],emp[8]))
            self.treeECalcEco.insert('','end',values=("'"+str(emp[1])+"'",emp[3],emp[4],emp[5],emp[6],emp[7],emp[8],emp[9]))

    def calcCoeficienteEva(self, emp):
        listPer = self.getPeriodo()
        sumeva = 0
        coeficiente = 0
        countDiv = len(listPer)
        for periodo in listPer:
            queryEva = "SELECT te.peso FROM postgres.public.tipo_evaluacion te \
            INNER JOIN postgres.public.evaluacion AS e ON te.id = e.evaluacion_tipoevaluacion_id WHERE e.evaluacion_perio_id="\
            +str(periodo[0])+" AND e.evaluacion_empleado_id = '"+emp+"'"
            self.cursorLoc.execute(queryEva)
            eva = self.cursorLoc.fetchone()
            if eva is not None:
                if eva[0] == 0:
                    countDiv -= 1
                else:
                    sumeva += eva[0]
            else:
                return messagebox.showinfo('Notificación',"'Trabajador '"+emp+"'No tiene registro de evaluación en el mes de "+str(periodo[1]))
        if countDiv > 0: 
            coeficiente = sumeva/countDiv
            if coeficiente >2:
                return coeficiente        
        return coeficiente

    def actualizartreeCALCECO1(self,event):
        self.actualizartreeCALCECO()  
    

    
    def cargarDpto(self):
        options=[]         
        queryP='SELECT x.* FROM postgres.public.area x order by area asc'
        self.cursorLoc.execute(queryP)
        slistArea=self.cursorLoc.fetchall()
        for row in slistArea:
            options.append(row[1])
        
        self.cb_area_calceco['values']=options

    def cargarEmpCB(self):
        options=[]         
        queryP='SELECT x.id FROM postgres.public.empleado x order by x.empleado_area_id asc'
        self.cursorLoc.execute(queryP)
        slistEmp=self.cursorLoc.fetchall()
        for row in slistEmp:
            options.append(row[0])
        
        self.cb_emp_ninterno['values']=options

    def getPeriodo(self):         
        queryP='SELECT p.* FROM postgres.public.utilidades_periodo_incluye x INNER JOIN postgres.public.periodo AS p ON x.upincluye_periodo_id = p.id order by p.id asc'
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchall()

    def getUtiliDist(self):
        queryUD="SELECT x.* FROM postgres.public.utilidades_distribucion x"
        self.cursorLoc.execute(queryUD)
        result = self.cursorLoc.fetchone()
        return result 


    def getMesesA(self):
        queryP="SELECT id_peri, nombre FROM ZUNpr.dbo.n_periodo WHERE id_peri >"+str(self.getPeriodo()[2][0]-1)        

        self.cursorZun.execute(queryP)
        listmeseA = self.cursorZun.fetchall() 
        options=[]
        for mes in listmeseA:
            options.append((str(mes['id_peri'])+'-'+str(mes['nombre'])))
        self.cb_periodo_calceco['values']=options

    def obtenerPerMes(self,mes):
        queryP="SELECT * FROM postgres.public.periodo where mes='"+str(mes)+"'"
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchall()      

    def getDepartamento(self,idemp):         
        queryP="SELECT a.area  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id where emp.id = "+str(idemp)
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchone()[0]


    #Mostrar reportes de otros pagos    
    def expDBF(self):                
        pass

    
    def getpagoSalMT(self, empleado):
        mtsalario = []
        querygetSal="SELECT ps.sal_devengado,ps.destajo,ps.horast,ps.psalario_periodo_id FROM postgres.public.utilidades_periodo_incluye up \
        INNER JOIN postgres.public.pago_salario AS ps ON up.upincluye_periodo_id = ps.psalario_periodo_id where ps.psalario_empleado_id  = '"+empleado+"' order by up.upincluye_periodo_id asc" 
        self.cursorLoc.execute(querygetSal)
        salarioList = self.cursorLoc.fetchall()
        for sal in salarioList:
            mtsalario.append(((sal[0]),sal[2],sal[1],sal[3]))
        return mtsalario
    

    def limpiarExcel(self,fila,url):         
        path = url
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet.delete_rows(fila, sheet.max_row-1)        
        wb.save(path)

    
