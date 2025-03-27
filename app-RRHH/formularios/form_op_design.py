import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox
from config import COLOR_CUERPO_PRINCIPAL, COLOR_BARRA_SUPERIOR,CONN_LOC,CURSOR_LOC
import openpyxl
import os
import subprocess
import tkinter.font as tkfont
from openpyxl.styles import Font, colors, fills, Alignment, PatternFill, NamedStyle



class FormularioOtrosPagosDesign():
    
    def __init__(self, panel_principal):   
       
        #Definiendo variables
        # Variables de conexion
        self.connLoc = CONN_LOC
        self.cursorLoc = CURSOR_LOC
        if self.getPeriodo():
            
            # Definiendo controles de seleccion
            self.store = self.cargarTP()
            self.empSelec = ''
            self.monto_str = StringVar()
            self.tx_empleado_op = ttk.Entry(panel_principal, font=('Times', 14), width=10)
            self.tx_empleado_op.grid(row=0,column=0,padx=5,pady=5,ipadx=40)

            #Boton para buscar empleados        
            self.btn_bempleados_op = tk.Button(panel_principal, text="Buscar", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.actualizartreeEOP)
            self.btn_bempleados_op.place(x=250, y=2)
            #Para buscar por departamento
            #Label area
            self.tx_area = tk.Label(panel_principal, font=('Times', 14), width=20, bg=COLOR_CUERPO_PRINCIPAL, text='Departamento:')
            self.tx_area.place(x=350, y=5)

            #Combo departamento
            self.cb_area_op= ttk.Combobox(panel_principal, width=30)
            self.cb_area_op.bind('<<ComboboxSelected>>', self.actualizartreeEOP1)
            #self.cb_periodo.current(0)
            self.cb_area_op.place(x=520, y=5)

            #Boton para Registrar pago        
            self.btn_agPago = tk.Button(panel_principal, text="Registrar pago", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.registrarOP)
            self.btn_agPago.place(x=720, y=290)     

            #Boton para Eliminar pago        
            self.btn_agPago = tk.Button(panel_principal, text="Eliminar pago", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.eliminarOP)
            self.btn_agPago.place(x=870, y=290)   
            
            


            #Boton Revisar otros pagos        
            self.btn_signEmpOP = tk.Button(panel_principal, text="Revisar otros pagos", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.signOP)
            self.btn_signEmpOP.place(x=720, y=350)

            #label empleado
            self.lb_sempleado_op = tk.Label(panel_principal, text='Empleado seleccionado', justify='right', bg=COLOR_CUERPO_PRINCIPAL, font=('Times', 12), width=30)
            self.lb_sempleado_op.place(x=720, y=100)       
            
            #label tipo pago
            self.lb_tp_op = tk.Label(panel_principal, text='Tipo de pago:', justify='right', bg=COLOR_CUERPO_PRINCIPAL, font=('Times', 12), width=12)
            self.lb_tp_op.place(x=710, y=151)
            #label monto
            self.lb_monto_op = tk.Label(panel_principal, text='Monto:', justify='right', bg=COLOR_CUERPO_PRINCIPAL, font=('Times', 12), width=10)
            self.lb_monto_op.place(x=700, y=251)           


            #ComboTipoPago
            self.cb_tp_op = ttk.Combobox(panel_principal, values = self.store, postcommand=self.cargarTP, width=21)
            #self.cb_tp_op.bind('<Configure>', self.on_combo_configure)
            self.cb_tp_op.place(x=808, y=150)

            #Periodo del pago
            self.lx_PPlabel = tk.Label(panel_principal, text="Mes:", justify='right', bg=COLOR_CUERPO_PRINCIPAL, font=('Times', 12))
            self.lx_PPlabel.place(x=720, y=201) 
            
            self.cb_periodo_op = ttk.Combobox(panel_principal, postcommand=self.cargarPeriodoOP, width=12)
            #self.cb_periodo.current(0)
            self.cb_periodo_op.place(x=808, y=200) 
            validatecommand = panel_principal.register(self.is_valid_char)
            self.tx_monto_op = ttk.Entry(panel_principal, validate="key", validatecommand=(validatecommand, "%S"),font=('Times', 14), width=10, textvariable=self.monto_str)
            #self.cb_periodo.current(0)
            self.tx_monto_op.place(x=808, y=251)
      
            
            

            #Treeview        
            columns = ('numero', 'nombreap', 'ci', 'area','opagos')
            self.treeEOP = ttk.Treeview(panel_principal, height=16, columns=columns, show='headings')
            self.style = ttk.Style(self.treeEOP)
            self.style.configure('Treeview',rowheight=30)
            self.treeEOP.column('numero',width=80)
            self.treeEOP.column('nombreap',width=200)
            self.treeEOP.column('ci',width=110)
            self.treeEOP.column('area',width=200)
            self.treeEOP.column('opagos',width=100)

            self.treeEOP.heading(column='numero', text='No.')
            self.treeEOP.heading(column='nombreap', text='Nombre y apellidos')
            self.treeEOP.heading(column='ci', text='CI')
            self.treeEOP.heading(column='area', text='Área')
            self.treeEOP.heading(column='opagos',text='Otros pagos')
            self.treeEOP.grid(row=1,column=0, columnspan=5,ipadx=5,padx=5,pady=5)
            self.treeEOP.bind('<<TreeviewSelect>>',self.selectEmp)
            self.actualizartreeEOP()  
              
            self.cargarDpto()   
        else:
            messagebox.showinfo('Notificación','Debe registrar un período de evaluación')

    def on_combo_configure(self,event):
        store = self.cargarTP()  
        maxlet = 0
        for el in store:
            if len(el) > maxlet:
                maxlet = len(el)
        #font = tkfont.nametofont(str(event.widget.cget('font')))
        width = maxlet*2
        #width = font.measure(store[0] + "0") - event.width
        style = ttk.Style()
        style.configure('TCombobox',(0,0,width,0))

    def is_valid_char(self,char):
        return char in "0123456789.-"

        
    #Listado de otros pagos
    def listOP(self, emp='', peri=''):
        if peri:
            queryLOP = "SELECT e.id, e.nombreap, e.ci, a.area, op.monto, tp.tipo FROM postgres.public.opago op \
                INNER JOIN postgres.public.empleado as e ON op.opago_empleado_id = e.id \
                    INNER JOIN postgres.public.tipo_pago AS tp ON op.tpago_id = tp.id INNER JOIN postgres.public.area AS a ON e.empleado_area_id = a.id WHERE e.id= "+str(emp)+" AND op.opago_periodo_id = "+str(peri)+" ORDER BY a.id"
        elif emp:
            queryLOP = "SELECT e.id, e.nombreap, e.ci, a.area, op.monto, tp.tipo FROM postgres.public.opago op \
                INNER JOIN postgres.public.empleado as e ON op.opago_empleado_id = e.id \
                    INNER JOIN postgres.public.tipo_pago AS tp ON op.tpago_id = tp.id INNER JOIN postgres.public.area AS a ON e.empleado_area_id = a.id WHERE e.id= "+str(emp)+" ORDER BY a.id"
        else:
            queryLOP = "SELECT e.id, e.nombreap, e.ci, a.area, op.monto, tp.tipo FROM postgres.public.opago op \
                INNER JOIN postgres.public.empleado as e ON op.opago_empleado_id = e.id \
                    INNER JOIN postgres.public.tipo_pago AS tp ON op.tpago_id = tp.id INNER JOIN postgres.public.area AS a ON e.empleado_area_id = a.id ORDER BY a.id"
        
        self.cursorLoc.execute(queryLOP)
        return self.cursorLoc.fetchall()

    #Cargar combo de periodo
    def cargarPeriodoOP(self): 
        slistp = self.getPeriodo() 
        options = []       
        for row in slistp:
            options.append(str(row[0])+"-"+str(row[1]))  

        self.cb_periodo_op['value']=options

    #Definiendo tree view de periodo
    def registrarOP(self):      
        if self.empSelec:
            idTPSelected = self.cb_tp_op.get().split('-')[0]
            idPeriodo = self.cb_periodo_op.get().split('-')[0]
            if self.tx_monto_op.get() != '' and  idPeriodo!= '' and  idTPSelected != '':
                selectedItem=self.treeEOP.item(self.empSelec)
                queryIOP = "INSERT INTO postgres.public.opago (tpago_id, monto, opago_periodo_id, opago_empleado_id) \
                    VALUES("+str(idTPSelected)+","+self.tx_monto_op.get()+","+str(idPeriodo)+","+str(selectedItem['values'][0])+")"
                self.cursorLoc.execute(queryIOP)
                self.connLoc.commit()            
                #self.treeEOP.set(self.empSelec, column='opagos', value=self.cb_tp_op.get())  
                messagebox.showinfo('Confirmación','La información del pago se registró satisfactoriamente') 
                self.cb_periodo_op.set('')
                self.cb_tp_op.set('')
                self.monto_str.set('')
                self.lb_sempleado_op['text']='Empleado seleccionado'
            else:
                messagebox.showinfo('Campos vacíos','Existen campos vacíos, debe completarlos')
                self.lb_sempleado_op['text']='Empleado seleccionado'
        else:            
            messagebox.showinfo('Información','Debe seleccionar un trabajador')
        self.actualizartreeEOP()

    def eliminarOP(self):
        if self.empSelec:
            idTPSelected = self.cb_tp_op.get().split('-')[0]
            idPeriodo = self.cb_periodo_op.get().split('-')[0]
            selectedItem=self.treeEOP.item(self.empSelec)
            cantOPEmpbefore = len(self.listOP(selectedItem['values'][0]))
            if self.tx_monto_op.get() != '' and  idPeriodo!= '' and  idTPSelected != '':                
                queryEOP = "DELETE FROM postgres.public.opago WHERE tpago_id = "+str(idTPSelected)+"\
                        AND monto = "+self.tx_monto_op.get()+" AND opago_periodo_id = "+str(idPeriodo)+" AND opago_empleado_id = "+str(selectedItem['values'][0])
                self.cursorLoc.execute(queryEOP)
                self.connLoc.commit() 
                cantP = len(self.listOP(selectedItem['values'][0]))
                if cantOPEmpbefore == cantP:
                    messagebox.showinfo('Sin acción','No existen registros para la información suministrada')
                    self.cb_periodo_op.set('')
                    self.cb_tp_op.set('')
                    self.monto_str.set('')
                    self.lb_sempleado_op['text']='Empleado seleccionado'
                    #self.treeEOP.set(self.empSelec, column='opagos', value=self.cb_tp_op.get())  
                else:
                    messagebox.showinfo('Confirmación','La información se eliminó correctamente') 
                    self.cb_periodo_op.set('')
                    self.cb_tp_op.set('')
                    self.monto_str.set('')
                    self.lb_sempleado_op['text']='Empleado seleccionado'
            else:
                messagebox.showinfo('Campos vacíos','Existen campos vacíos, debe completarlos')
                self.lb_sempleado_op['text']='Empleado seleccionado'
        else:            
            messagebox.showinfo('Información','Debe seleccionar un trabajador')
        self.actualizartreeEOP()


        
    def actualizartreeEOP(self):
        self.treeEOP.delete(*self.treeEOP.get_children())         
        queryEmpL=''
        if self.tx_empleado_op.get() != '' and self.cb_area_op.get() == '':
            queryEmpL="SELECT x.id,x.nombreap,x.ci,a.area FROM postgres.public.empleado AS x INNER JOIN postgres.public.area AS a ON x.empleado_area_id = a.id where x.nombreap like '%"+self.tx_empleado_op.get().upper()+"%' ORDER BY a.id ASC"
        elif self.cb_area_op.get() != '' and self.tx_empleado_op.get() == '':
            queryEmpL="SELECT x.id,x.nombreap,x.ci,a.area FROM postgres.public.empleado AS x INNER JOIN postgres.public.area AS a ON x.empleado_area_id = a.id where a.area = '"+self.cb_area_op.get()+"' ORDER BY a.id ASC"
        elif self.tx_empleado_op.get() != '' and self.cb_area_op.get() != '':
            queryEmpL="SELECT x.id,x.nombreap,x.ci,a.area FROM postgres.public.empleado AS x INNER JOIN postgres.public.area AS a ON x.empleado_area_id = a.id where x.nombreap like '%"+self.tx_empleado_op.get().upper()+"%' and a.area = '"+self.cb_area_op.get()+"' ORDER BY a.id ASC"
        else:
            queryEmpL='SELECT x.id,x.nombreap,x.ci,a.area FROM postgres.public.empleado AS x INNER JOIN postgres.public.area AS a ON x.empleado_area_id = a.id  ORDER BY a.id'
        
         
        self.cursorLoc.execute(queryEmpL)

        slistEmp = self.cursorLoc.fetchall()            
        for emp in slistEmp:
            if self.listOP("'"+emp[0]+"'") is not None:
                cantP = len(self.listOP("'"+emp[0]+"'"))
                self.treeEOP.insert('','end',values=("'"+emp[0]+"'",emp[1],emp[2],emp[3],cantP)) 
            else:
                self.treeEOP.insert('','end',values=("'"+emp[0]+"'",emp[1],emp[2],emp[3],'0'))


    def actualizartreeEOP1(self,event):
        self.actualizartreeEOP()

    def selectEmp(self,event):
        self.empSelec = self.treeEOP.selection()
        selectItem=self.treeEOP.item(self.empSelec)
        cadena=str(selectItem['values'][0])+" "+str(selectItem['values'][1])
        if len(cadena) < 26:
            self.lb_sempleado_op['text']=cadena
        else:
            self.lb_sempleado_op['text']=cadena[0:25]+"..."
    

    def cargarTP(self):
        options=[]         
        queryP='SELECT x.* FROM postgres.public.tipo_pago x order by id asc'
        self.cursorLoc.execute(queryP)
        slistTP=self.cursorLoc.fetchall()
        for row in slistTP:
            options.append(str(row[0])+'-'+row[1])        
        return options
    
    def cargarDpto(self):
        options=[]         
        queryP='SELECT x.* FROM postgres.public.area x order by area asc'
        self.cursorLoc.execute(queryP)
        slistArea=self.cursorLoc.fetchall()
        for row in slistArea:
            options.append(row[1])
        
        self.cb_area_op['values']=options

    def getPeriodo(self):         
        queryP='SELECT p.* FROM postgres.public.utilidades_periodo_incluye x INNER JOIN postgres.public.periodo AS p ON x.upincluye_periodo_id = p.id order by p.id asc'
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchall()

    def obtenerPerMes(self,mes):
        queryP="SELECT * FROM postgres.public.periodo where mes='"+str(mes)+"'"
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchall()      

    def getDepartamento(self,idemp):         
        queryP="SELECT a.area  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id where emp.id = "+str(idemp)
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchone()[0]

    #quitar merge a las celdas del reporte de op
    def unmergexlsop(self):
        path = "file/list_opagos.xlsx"
        self.limpiarExcel(4,path) 
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        for items in sorted(sheet.merged_cell_ranges):
            if str(items) != 'A1:I2':
                sheet.unmerge_cells(str(items))

        wb.save(path)


    #Mostrar reportes de otros pagos    
    def signOP(self): 
               
        path = "file/list_opagos.xlsx"
        self.limpiarExcel(4,path) 
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        consec = 1
        row = 4
        text_format = Font(
            bold = True,
            name = 'Calibri',
            size = '-1',
            color = colors.BLACK )   
        sheet['G3']=self.getPeriodo()[0][1].upper()
        sheet['G3'].font += text_format 
        sheet['H3']=self.getPeriodo()[1][1].upper()
        sheet['H3'].font += text_format
        sheet['I3']=self.getPeriodo()[2][1].upper()
        sheet['I3'].font += text_format
        
        
        queryP="SELECT a.area,emp.id,emp.ci,emp.nombreap,emp.escalas,emp.thoraria,emp.destajo  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id ORDER BY a.id"
        self.cursorLoc.execute(queryP)
        listEmp = self.cursorLoc.fetchall()
        for empleado in listEmp:
            countmerge = len(self.listOP("'"+empleado[1]+"'"))
            sheet["A"+str(row)].alignment = Alignment(vertical='top')
            sheet["B"+str(row)].alignment = Alignment(vertical='top')
            sheet["C"+str(row)].alignment = Alignment(vertical='top')
            sheet["D"+str(row)].alignment = Alignment(vertical='top')
            sheet["E"+str(row)].alignment = Alignment(vertical='top')         
            
            
            if len(self.listOP("'"+empleado[1]+"'")) > 0:
                
                if countmerge > 1:
                    sheet.merge_cells("A"+str(row)+":A"+str(row+countmerge-1))                     
                    sheet.merge_cells("B"+str(row)+":B"+str(row+countmerge-1))                    
                    sheet.merge_cells("C"+str(row)+":C"+str(row+countmerge-1))                    
                    sheet.merge_cells("D"+str(row)+":D"+str(row+countmerge-1))                    
                    sheet.merge_cells("E"+str(row)+":E"+str(row+countmerge-1))                    

                sheet['A'+str(row)]=consec
                sheet['B'+str(row)]=empleado[1]
                sheet['C'+str(row)]=empleado[3]
                sheet['D'+str(row)]=empleado[2]
                sheet['E'+str(row)]=empleado[0]
                if self.listOP("'"+empleado[1]+"'",self.getPeriodo()[0][0]) is not None:
                    listop = self.listOP("'"+empleado[1]+"'",self.getPeriodo()[0][0])
                    for p in listop:
                        sheet['F'+str(row)]=p[5]   
                        sheet["F"+str(row)].alignment = Alignment(vertical='top', wrapText=True)                    
                        sheet['G'+str(row)]=p[4]
                        sheet["G"+str(row)].alignment = Alignment(vertical='top', horizontal = 'center')
                        row+=1
                if self.listOP("'"+empleado[1]+"'",self.getPeriodo()[1][0]) is not None:
                    listop = self.listOP("'"+empleado[1]+"'",self.getPeriodo()[1][0])
                    for p in listop:
                        sheet['F'+str(row)]=p[5] 
                        sheet["F"+str(row)].alignment = Alignment(vertical='top', wrapText=True)                       
                        sheet['H'+str(row)]=p[4]
                        sheet["H"+str(row)].alignment = Alignment(vertical='top', horizontal = 'center')
                        row+=1
                if self.listOP("'"+empleado[1]+"'",self.getPeriodo()[2][0]) is not None:
                    listop = self.listOP("'"+empleado[1]+"'",self.getPeriodo()[2][0])
                    for p in listop:
                        sheet['F'+str(row)]=p[5]
                        sheet["F"+str(row)].alignment = Alignment(vertical='top', wrapText=True)
                        sheet['I'+str(row)]=p[4]
                        sheet["I"+str(row)].alignment = Alignment(vertical='top', horizontal = 'center')
                        row+=1
                
                consec+=1
            #Insertar empleados
            
            
        wb.save(path)
        self.convert_xlsx_to_pdf(path,'list_opagos')
        self.unmergexlsop()

    def convert_xlsx_to_pdf(self,xlsx_file,nombreA=''):
        try:
            subprocess.run(["libreoffice24.2", "--headless", "--convert-to", "pdf", xlsx_file])
            separador = os.path.sep
            dir_actual = os.path.dirname(os.path.abspath(__file__))
            dir = separador.join(dir_actual.split(separador)[:-1])
            #dirfile = separador.join(xlsx_file.split(separador))
            command =  ['open', dir+separador+nombreA+'.pdf']
            subprocess.run(command,shell=False)
            print("Done!")

        except Exception as e:
            print("Error:", e)

    

    def limpiarExcel(self,fila,url):         
        path = url
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet.delete_rows(fila, sheet.max_row-1)        
        wb.save(path)

    
