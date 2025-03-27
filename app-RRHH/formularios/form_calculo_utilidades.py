import tkinter as tk
from tkinter import *
from decimal import *
from tkinter import ttk, messagebox
from config import COLOR_CUERPO_PRINCIPAL, COLOR_BARRA_SUPERIOR, CONN_ZUN,CURSOR_ZUN,CONN_LOC,CURSOR_LOC
from PIL import Image, ImageTk
import util.util_imagenes as util_img
import openpyxl
from openpyxl.styles import Font, colors, fills, Alignment, PatternFill, NamedStyle
import subprocess
import os



class FormularioCalcUtilidadesDesign():

    def __init__(self, panel_principal):   
       
        #Definiendo variables
        # Variables de conexion
        #Conexion
        self.connZun = CONN_ZUN
        self.cursorZun = CURSOR_ZUN

        self.connLoc = CONN_LOC
        self.cursorLoc = CURSOR_LOC
        if self.getPeriodo():
            validatecommand = panel_principal.register(self.is_valid_char)
            #variablesde estado
            self.registro_salario = False
            self.registro_vacaciones = False
            self.calculo_utili = False
            # Definiendo controles de seleccion
            self.empSelec = ''
            self.inv=[]
            self.inv_str = StringVar()
            self.tx_empleado = ttk.Entry(panel_principal, font=('Times', 14), width=10)
            self.tx_empleado.grid(row=0,column=0,padx=5,pady=5,ipadx=40)

            # Definiendo monto de distribucion de utilidades
            self.lb_monto = tk.Label(panel_principal, font=('Times', 12), bg=COLOR_CUERPO_PRINCIPAL, text='Monto a distribuir:')
            self.lb_monto.place(x=750, y=400)

            self.tx_distribuir = ttk.Entry(panel_principal, validate="key", validatecommand=(validatecommand, "%S"), font=('Times', 14), width=13)
            self.tx_distribuir.place(x=875, y=400) 

            self.btn_utilidades = tk.Button(panel_principal, text="Reporte de utilidades", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.distribuirUtil)
            self.btn_utilidades.place(x=750, y=490)

            #Boton para buscar empleados        
            self.btn_bempleados = tk.Button(panel_principal, text="Buscar", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.actualizartreeEUtil)
            self.btn_bempleados.place(x=250, y=2)
            
            
            #Label mostrasr total de registros
            self.tx_total = tk.Label(panel_principal, font=('Times', 18), bg=COLOR_CUERPO_PRINCIPAL, text='Total de registros: 0')
            self.tx_total.place(x=750, y=120)

            #Label mostrasr total de registros
            self.tx_diferencia = tk.Label(panel_principal, font=('Times', 12), bg=COLOR_CUERPO_PRINCIPAL, text='Diferencia: 0')
            self.tx_diferencia.place(x=750, y=425)
            
            #Para buscar por departamento
            #Label departamento
            self.tx_departamento = tk.Label(panel_principal, font=('Times', 14), width=20, bg=COLOR_CUERPO_PRINCIPAL, text='Departamento:')
            self.tx_departamento.place(x=350, y=5)

            #Combo departamento
            self.cb_departamento= ttk.Combobox(panel_principal, width=30)
            self.cb_departamento.bind('<<ComboboxSelected>>', self.actualizartreeEUtil1)
            #self.cb_periodo.current(0)
            self.cb_departamento.place(x=520, y=5)

            #Boton para registrar salario       
            self.btn_agSal = tk.Button(panel_principal, text="Registrar salario", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.regSal)
            self.btn_agSal.place(x=750, y=80)

            #Boton para registrar vacaciones        
            self.btn_agVaca = tk.Button(panel_principal, text="Registrar vacaciones", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.regVac)
            self.btn_agVaca.place(x=750, y=160)            

            style = ttk.Style()            
            style.configure('TLabelframe', background=COLOR_CUERPO_PRINCIPAL, borderwidth=2, bodercolor='black')
            style.configure('TLabelframe.Label', background=COLOR_CUERPO_PRINCIPAL)

            #Label frame para las invalidadnte
            self.lb_frame = ttk.Labelframe(panel_principal, text='Invalidantes del pago', style='TLabelframe')
            self.lb_frame.place(x=735, y=200, width=255,height=195,)

            #Empleado seleccionado
            self.lb_sempleado_cu = tk.Label(self.lb_frame, text='Empleado:',width=10, justify='center', bg=COLOR_CUERPO_PRINCIPAL, font=('Times', 11))
            self.lb_sempleado_cu.grid(row=0,column=0)
            self.lb_sempleado_cu.grid_propagate(False)

            #Listado de empleados por id
            self.cb_empleado= ttk.Combobox(self.lb_frame, width=10)
            #self.cb_periodo.current(0)
            self.cb_empleado.grid(row=0,column=1, padx=5, pady=5)

            #Descripción invalidantes
            self.textoComentInv=tk.Text(self.lb_frame, width=30, height=6, font=('Times', 11))
            #self.textoComentInv.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
            self.textoComentInv.grid(row=1,column=0,columnspan=3,padx=5,pady=5)
            self.scrollVert=ttk.Scrollbar(self.lb_frame, command=self.textoComentInv.yview)
            #self.scrollVert.pack(fill=tk.Y, side=tk.RIGHT) 
            self.scrollVert.grid(row=1,column=3, sticky = tk.NS)

            #Definir imagenes de botones
            imagen_pil_btadd = Image.open("./imagenes/add.png")
            imagen_pil_btadd = imagen_pil_btadd.resize((20,20))
            imagen_btadd_tk = ImageTk.PhotoImage(imagen_pil_btadd)

            imagen_pil_btdel = Image.open("./imagenes/delete.png")
            imagen_pil_btdel = imagen_pil_btdel.resize((20,20))
            imagen_btdel_tk = ImageTk.PhotoImage(imagen_pil_btdel)

            imagen_pil_btlist = Image.open("./imagenes/list.png")
            imagen_pil_btlist = imagen_pil_btlist.resize((20,20))
            imagen_btlist_tk = ImageTk.PhotoImage(imagen_pil_btlist)

            #Boton list inv        
            self.btn_listinv = tk.Button(self.lb_frame, text="\uf0c9",bd=0, image=imagen_btlist_tk, font=(
                'Times', 13), command=self.listInv)
            self.btn_listinv.image=imagen_btlist_tk
            self.btn_listinv.grid(row=2,column=2)

            #Boton add inv        
            self.btn_addinv = tk.Button(self.lb_frame, text="\uf0c9",bd=0, image=imagen_btadd_tk, font=(
                'Times', 13), command=self.addInv)
            self.btn_addinv.image=imagen_btadd_tk
            self.btn_addinv.grid(row=2,column=0)

            #Boton elim inv        
            self.btn_elimInv = tk.Button(self.lb_frame, text="\uf0c9",bd=0, image=imagen_btdel_tk, font=(
                'Times', 13),  command=self.deleteInv)
            self.btn_elimInv.image=imagen_btdel_tk
            self.btn_elimInv.grid(row=2,column=1)      

            #Boton mostrar resumen        
            self.btn_showResume = tk.Button(panel_principal, text="Calcular resumen", font=(
                'Times', 13), bg=COLOR_BARRA_SUPERIOR, bd=0, fg=COLOR_CUERPO_PRINCIPAL, command=self.showResumen)
            self.btn_showResume.place(x=750, y=450)                 
            

            #Treeview        
            columns = ('numero', 'nombreap', 'ci', 'salario','vacaciones','horast','coef','devengado','descrinv')
            self.treeEUtil = ttk.Treeview(panel_principal, height=16, columns=columns, show='headings')
            self.treeEUtil["displaycolumns"]=('numero', 'nombreap', 'ci', 'salario','vacaciones','horast','coef','devengado')
            self.style = ttk.Style(self.treeEUtil)
            self.style.configure('Treeview',rowheight=30)
            self.treeEUtil.column('numero',width=80)
            self.treeEUtil.column('nombreap',width=200)
            self.treeEUtil.column('ci',width=110)
            self.treeEUtil.column('salario',width=60)
            self.treeEUtil.column('vacaciones',width=60)
            self.treeEUtil.column('horast',width=60)
            self.treeEUtil.column('coef',width=60)
            self.treeEUtil.column('devengado',width=80)

            self.treeEUtil.heading(column='numero', text='No.')
            self.treeEUtil.heading(column='nombreap', text='Nombre y apellidos')
            self.treeEUtil.heading(column='ci', text='CI')
            self.treeEUtil.heading(column='salario', text='Mt. Sal')
            self.treeEUtil.heading(column='vacaciones', text='Mt. Vac')
            self.treeEUtil.heading(column='horast', text='Horas T.')
            self.treeEUtil.heading(column='coef', text='C. Eva')
            self.treeEUtil.heading(column='devengado', text='S. Dev')
            self.treeEUtil.grid(row=1,column=0, columnspan=5,ipadx=5,padx=5,pady=5)
            self.treeEUtil.bind('<<TreeviewSelect>>',self.selectEmp)
            self.actualizartreeEUtil() 
            self.cargarEmpCB()
            self.cargarDpto() 
            
            #print(self.getVacacionesMT('0091'))
        else:
            messagebox.showinfo('Notificación','Debe registrar un período de evaluación')

    def getPagoDestajoZun(self,emp,periodo):
        tdestajo = 0
        queryDestajo = "SELECT hp.importe FROM ZUNpr.dbo.h_incidencias AS hi INNER JOIN ZUNpr.dbo.h_pagos AS hp ON  hi.id_inci = hp.id_inci \
            WHERE hp.id_cpago = 30 AND hi.no_interno = '"+emp+"' AND hi.id_ppago = "+str(periodo)  
        #print(queryDestajo)
        self.cursorZun.execute(queryDestajo)
        listDest=self.cursorZun.fetchall()
        for dest in listDest:
            tdestajo += dest['importe']
        #print(tdestajo)
        return tdestajo

    def is_valid_char(self,char):
        return char in "0123456789."

    def cargarEmpCB(self):
        options=[]         
        queryP='SELECT x.id FROM postgres.public.empleado x order by x.id asc'
        self.cursorLoc.execute(queryP)
        slistEmp=self.cursorLoc.fetchall()
        for row in slistEmp:
            options.append(row[0])
        
        self.cb_empleado['values']=options

    def deleteInv(self):
        if self.cb_empleado.get():           
            idemp=self.cb_empleado.get()
            if self.inv != []:
                for einv in self.inv:
                    finded = False
                    if einv[0]==idemp:
                        self.inv.remove(einv)
                        finded = True
                        messagebox.showinfo('Confirmación','Se eliminó el trabajador invalidado correctamente')
                        self.treeEUtil.selection_remove(self.treeEUtil.selection())
                        self.textoComentInv.delete(1.0, END)
                        self.cb_empleado.set('')
                if finded != True:
                    messagebox.showinfo('Sin acción','No existe ese trabajador en el registro')
            else:
                messagebox.showinfo('Error de validación','No existen elementos para eliminar')
        else:
            messagebox.showinfo('Información','Debe seleccionar el trabajador')
        


    def addInv(self):
        if self.cb_empleado.get():            
            idemp=self.cb_empleado.get()
            textarea = self.textoComentInv.get("1.0", tk.END)
            if textarea != '\n':
                finded = False
                for einv in self.inv:
                    if einv[0] == idemp:
                        finded = True
                if finded:
                    return messagebox.showwarning('Registro repetido','El especialista ya se encentra invalidado')
                else:                    
                    self.inv.append((idemp,textarea))
                    messagebox.showinfo('Confirmación','Se registró el trabajador invalidado correctamente')
                    self.treeEUtil.selection_remove(self.treeEUtil.selection())                    
                    self.textoComentInv.delete("1.0", END)
                    self.cb_empleado.set('')
            else:
                messagebox.showerror('Error de validación','Debe teclear una descripción')
        else:
            messagebox.showinfo('Información','Debe seleccionar el trabajador')

    def listInv(self):
        path = "file/list_invalidados.xlsx"
        row = 4 
        controw = 1
        self.limpiarExcel(row,path)   
        wb = openpyxl.load_workbook(path)
        sheet = wb.active  
        queryP="SELECT a.area,emp.id,emp.ci,emp.nombreap,emp.escalas,emp.thoraria  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id ORDER BY a.id"
        self.cursorLoc.execute(queryP)
        listEmp = self.cursorLoc.fetchall()
        for empleado in listEmp:
            for t in self.inv:
                idemp= empleado[1]
                if idemp in t:
                    sheet['A'+str(row)] = controw
                    sheet['B'+str(row)]=empleado[1]
                    sheet['B'+str(row)].alignment = Alignment(vertical='top', horizontal = 'left')
                    sheet['C'+str(row)]=empleado[3]
                    sheet['C'+str(row)].alignment = Alignment(vertical='top', horizontal = 'left')
                    sheet['D'+str(row)]=empleado[2]
                    sheet['D'+str(row)].alignment = Alignment(vertical='top', horizontal = 'left')
                    sheet['E'+str(row)]=empleado[0]   
                    sheet['E'+str(row)].alignment = Alignment(vertical='top', horizontal = 'left')                 
                    cadena = t[1]
                    sheet['F'+str(row)] = cadena.strip()
                    sheet['F'+str(row)].alignment = Alignment(vertical='top', wrapText=True)
                    controw +=1
                    row+=1
        wb.save(path)

        self.convert_xlsx_to_pdf(path,"list_invalidados")
        


    #Definiendo registro de salario
    def regSal(self):                
        queryEmpL=''
        countReg = 0
        self.limpiarNominaLoc()
        queryEmpL='SELECT id FROM postgres.public.empleado'
        
        self.cursorLoc.execute(queryEmpL)
        slistEmp = self.cursorLoc.fetchall()            
        for row in slistEmp:
            #Cargar nomina del primer mes del periodo
            queryGetNom1 = "SELECT x.Id_nomina_sal,x.no_interno,x.deveng_salario,x.dias_lab FROM ZUNpr.dbo.nomina_sal x where x.no_interno="+row[0]+" and x.id_periodo="+str(self.getPeriodo()[0][0])
            self.cursorZun.execute(queryGetNom1)
            sal_emp=self.cursorZun.fetchone()
            if sal_emp is not None:
                destajo = self.getPagoDestajoZun(row[0],self.getPeriodo()[0][0])            
                queryInsertSalLoc = "INSERT INTO postgres.public.pago_salario (id, sal_devengado,	destajo,	horast,	psalario_empleado_id,	psalario_periodo_id) VALUES("+str(sal_emp['Id_nomina_sal'])+","+str(sal_emp['deveng_salario'])+","+str(destajo)+","+str(sal_emp['dias_lab'])+",'"+row[0]+"',"+str(self.getPeriodo()[0][0])+")"
                self.cursorLoc.execute(queryInsertSalLoc)
                self.connLoc.commit()
            #Cargar nomina del segundo mes del periodo
            queryGetNom1 = "SELECT x.Id_nomina_sal,x.no_interno,x.deveng_salario,x.dias_lab FROM ZUNpr.dbo.nomina_sal x where x.no_interno="+row[0]+" and x.id_periodo="+str(self.getPeriodo()[1][0])
            self.cursorZun.execute(queryGetNom1)
            sal_emp=self.cursorZun.fetchone()
            if sal_emp is not None:
                destajo = self.getPagoDestajoZun(row[0],self.getPeriodo()[1][0]) 
                queryInsertSalLoc = "INSERT INTO postgres.public.pago_salario (id, sal_devengado,	destajo,	horast,	psalario_empleado_id,	psalario_periodo_id) VALUES("+str(sal_emp['Id_nomina_sal'])+","+str(sal_emp['deveng_salario'])+","+str(destajo)+","+str(sal_emp['dias_lab'])+",'"+row[0]+"',"+str(self.getPeriodo()[1][0])+")"
                self.cursorLoc.execute(queryInsertSalLoc)
                self.connLoc.commit()
            #Cargar nomina del tercer mes del periodo
            queryGetNom1 = "SELECT x.Id_nomina_sal,x.no_interno,x.deveng_salario,x.dias_lab FROM ZUNpr.dbo.nomina_sal x where x.no_interno="+row[0]+" and x.id_periodo="+str(self.getPeriodo()[2][0])
            self.cursorZun.execute(queryGetNom1)
            sal_emp=self.cursorZun.fetchone()
            if sal_emp is not None:
                destajo = self.getPagoDestajoZun(row[0],self.getPeriodo()[2][0])
                queryInsertSalLoc = "INSERT INTO postgres.public.pago_salario (id, sal_devengado,	destajo,	horast,	psalario_empleado_id,	psalario_periodo_id) VALUES("+str(sal_emp['Id_nomina_sal'])+","+str(sal_emp['deveng_salario'])+","+str(destajo)+","+str(sal_emp['dias_lab'])+",'"+row[0]+"',"+str(self.getPeriodo()[2][0])+")"
                self.cursorLoc.execute(queryInsertSalLoc)
                self.connLoc.commit()
            countReg+=1
        self.tx_total['text'] = 'Total de registros: '+str(countReg)
        self.registro_salario = True
        messagebox.showinfo('Confirmación','La información de la nómina de salario se registró satisfactoriamente')

    def regVac(self):               
        queryEmpL='SELECT id FROM postgres.public.empleado'
        self.limpiarVacacionesLoc()
        self.cursorLoc.execute(queryEmpL)
        slistEmp = self.cursorLoc.fetchall()            
        for row in slistEmp:
            #Cargar vacaciones del mes-1 del periodo            
            id_inci=self.getVacaInciTrab(row[0],self.getPeriodo1()[0][0]) 
            if id_inci is not None:            
                queryVaca = "SELECT v.id_inci,v.tiempo_total,v.importe_total,v.dias_periodo,importe_periodo FROM ZUNpr.dbo.h_vacaciones v \
                WHERE v.id_inci ="+str(id_inci['id_inci'])  
                self.cursorZun.execute(queryVaca)
                dataVacaciones = self.cursorZun.fetchone()
                if dataVacaciones['dias_periodo'] == 0:
                    queryInsertVacaLoc = "INSERT INTO postgres.public.vacacionesp (id,dias,monto,vacacionesp_empleado_id,vacacionesp_periodo_id,tiempo_tota,importe_total) \
                    VALUES("+str(id_inci['id_inci'])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+",'"+row[0]+"',"+str(self.getPeriodo1()[0][0])+","+str(dataVacaciones['tiempo_total'])+","+str(dataVacaciones['importe_total'])+")"
                    self.cursorLoc.execute(queryInsertVacaLoc)
                    self.connLoc.commit()
                if dataVacaciones['importe_total'] > dataVacaciones['importe_periodo'] and dataVacaciones['dias_periodo'] != 0:
                    importe_total = dataVacaciones['importe_total'] - dataVacaciones['importe_periodo']
                    tiempo_total = dataVacaciones['tiempo_total'] - dataVacaciones['dias_periodo']
                    queryInsertVacaLoc = "INSERT INTO postgres.public.vacacionesp (id,dias,monto,vacacionesp_empleado_id,vacacionesp_periodo_id,tiempo_tota,importe_total) \
                    VALUES("+str(id_inci['id_inci'])+",0,0,'"+row[0]+"',"+str(self.getPeriodo1()[0][0])+","+str(tiempo_total)+","+str(importe_total)+")"
                    self.cursorLoc.execute(queryInsertVacaLoc)
                    self.connLoc.commit()
                
            #Cargar vacaciones del mes 1 del periodo            
            id_inci=self.getVacaInciTrab(row[0],(self.getPeriodo1()[1][0]))                      
            if id_inci is not None:            
                queryVaca = "SELECT v.id_inci,v.tiempo_total,v.importe_total,v.dias_periodo,importe_periodo FROM ZUNpr.dbo.h_vacaciones v \
                WHERE v.id_inci ="+str(id_inci['id_inci'])
                self.cursorZun.execute(queryVaca)
                dataVacaciones = self.cursorZun.fetchone()
                if dataVacaciones['tiempo_total'] == 0:
                    queryInsertVacaLoc = "INSERT INTO postgres.public.vacacionesp (id,dias,monto,vacacionesp_empleado_id,vacacionesp_periodo_id,tiempo_tota,importe_total) \
                    VALUES("+str(id_inci['id_inci'])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+",'"+row[0]+"',"+str(self.getPeriodo1()[1][0])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+")"
                    self.cursorLoc.execute(queryInsertVacaLoc)
                    self.connLoc.commit()
                else:
                    queryInsertVacaLoc = "INSERT INTO postgres.public.vacacionesp (id,dias,monto,vacacionesp_empleado_id,vacacionesp_periodo_id,tiempo_tota,importe_total) \
                    VALUES("+str(id_inci['id_inci'])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+",'"+row[0]+"',"+str(self.getPeriodo1()[1][0])+","+str(dataVacaciones['tiempo_total'])+","+str(dataVacaciones['importe_total'])+")"
                    self.cursorLoc.execute(queryInsertVacaLoc)
                    self.connLoc.commit()
            #Cargar vacaciones del mes 2 del periodo            
            id_inci=self.getVacaInciTrab(row[0],(self.getPeriodo1()[2][0]))            
            if id_inci is not None:   
                queryVaca = "SELECT v.id_inci,v.tiempo_total,v.importe_total,v.dias_periodo,importe_periodo FROM ZUNpr.dbo.h_vacaciones v \
                WHERE v.id_inci ="+str(id_inci['id_inci'])
                self.cursorZun.execute(queryVaca)
                dataVacaciones = self.cursorZun.fetchone()         
                if dataVacaciones['tiempo_total'] == 0:
                    queryInsertVacaLoc = "INSERT INTO postgres.public.vacacionesp (id,dias,monto,vacacionesp_empleado_id,vacacionesp_periodo_id,tiempo_tota,importe_total) \
                    VALUES("+str(id_inci['id_inci'])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+",'"+row[0]+"',"+str(self.getPeriodo1()[2][0])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+")"
                    self.cursorLoc.execute(queryInsertVacaLoc)
                    self.connLoc.commit()
                else:
                    queryInsertVacaLoc = "INSERT INTO postgres.public.vacacionesp (id,dias,monto,vacacionesp_empleado_id,vacacionesp_periodo_id,tiempo_tota,importe_total) \
                    VALUES("+str(id_inci['id_inci'])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+",'"+row[0]+"',"+str(self.getPeriodo1()[2][0])+","+str(dataVacaciones['tiempo_total'])+","+str(dataVacaciones['importe_total'])+")"
                    self.cursorLoc.execute(queryInsertVacaLoc)
                    self.connLoc.commit()
            #Cargar vacaciones del mes 3 del periodo            
            id_inci=self.getVacaInciTrab(row[0],(self.getPeriodo1()[3][0]))            
            if id_inci is not None:                       
                queryVaca = "SELECT v.id_inci,v.tiempo_total,v.importe_total,v.dias_periodo,importe_periodo FROM ZUNpr.dbo.h_vacaciones v \
                WHERE v.id_inci ="+str(id_inci['id_inci'])
                self.cursorZun.execute(queryVaca)
                dataVacaciones = self.cursorZun.fetchone()
                if dataVacaciones['dias_periodo'] != 0:
                    queryInsertVacaLoc = "INSERT INTO postgres.public.vacacionesp (id,dias,monto,vacacionesp_empleado_id,vacacionesp_periodo_id,tiempo_tota,importe_total) \
                    VALUES("+str(id_inci['id_inci'])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+",'"+row[0]+"',"+str(self.getPeriodo1()[3][0])+","+str(dataVacaciones['dias_periodo'])+","+str(dataVacaciones['importe_periodo'])+")"
                    self.cursorLoc.execute(queryInsertVacaLoc)
                    self.connLoc.commit()
        self.registro_vacaciones = True
        messagebox.showinfo('Confirmación','La información de las vacaciones se registró satisfactoriamente')

    def selectEmp(self,event):
        self.empSelec = self.treeEUtil.selection()
        selectItem=self.treeEUtil.item(self.empSelec)               
        self.textoComentInv.insert(1.0,selectItem['values'][8])
        self.cb_empleado.set(selectItem['values'][0].replace("'",""))

    def distribuirUtil(self):
        if self.calculo_utili == True:
            path = "file/utilidades_dist.xlsx"
            row = 6 
            controw = 1
            self.limpiarExcel(row,path)   
            wb = openpyxl.load_workbook(path)
            sheet = wb.active
            montoDistribuir = self.getUtiliDist()[2]#Decimal(self.tx_distribuir.get())
            sheet['I3'] = montoDistribuir
            sheet['I3'].number_format = '#,##0.00'
            
            alignmentText = Alignment(horizontal=LEFT)
            alignmentNumber = Alignment(horizontal=CENTER)
            text_format = Font(
            bold = False,
            name = 'Calibri',
            size = '0',
            color = colors.BLACK )   
            number_format = Font(
            bold = False,
            name = 'Calibri',
            size = '0',
            color = colors.BLACK) 
            sheet['N5'] =  self.getPeriodo()[0][1]
            sheet['O5'] =  self.getPeriodo()[1][1]
            sheet['P5'] =  self.getPeriodo()[2][1]

            sheet['G4'] =  self.getPeriodo()[0][1]
            sheet['I4'] =  self.getPeriodo()[1][1]
            sheet['K4'] =  self.getPeriodo()[2][1]

            queryP="SELECT a.area,emp.id,emp.ci,emp.nombreap,emp.escalas,emp.thoraria  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id ORDER BY a.id"
            self.cursorLoc.execute(queryP)
            listEmp = self.cursorLoc.fetchall()
            for empleado in listEmp:
                

                sheet['B'+str(row)] =  controw
                sheet['B'+str(row)].font +=  number_format
                sheet['B'+str(row)].alignment += alignmentNumber
                sheet['C'+str(row)] =  empleado[1]
                sheet['C'+str(row)].font +=  text_format
                sheet['C'+str(row)].alignment += alignmentText
                sheet['D'+str(row)] =  empleado[2]
                sheet['D'+str(row)].font +=  text_format
                sheet['D'+str(row)].alignment += alignmentText
                sheet['E'+str(row)] =  empleado[3]
                sheet['E'+str(row)].font +=  text_format
                sheet['E'+str(row)].alignment += alignmentText
                sheet['F'+str(row)] =  empleado[4]
                sheet['F'+str(row)].font +=  text_format
                sheet['F'+str(row)].alignment += alignmentText

                    
                sheet['G'+str(row)] =  '0'
                sheet['G'+str(row)].font +=  number_format
                sheet['G'+str(row)].alignment += alignmentNumber                
                sheet['H'+str(row)] =  '0'
                sheet['H'+str(row)].font +=  number_format
                sheet['H'+str(row)].alignment += alignmentNumber
                sheet['I'+str(row)] =  '0'
                sheet['I'+str(row)].font +=  number_format
                sheet['I'+str(row)].alignment += alignmentNumber
                sheet['J'+str(row)] =  '0'
                sheet['J'+str(row)].font +=  number_format
                sheet['J'+str(row)].alignment += alignmentNumber
                sheet['K'+str(row)] =  '0'
                sheet['K'+str(row)].font +=  number_format
                sheet['K'+str(row)].alignment += alignmentNumber
                sheet['L'+str(row)] =  'NE'
                sheet['L'+str(row)].font +=  number_format
                sheet['L'+str(row)].alignment += alignmentNumber
                sheet['M'+str(row)] =  'NE'
                sheet['M'+str(row)].font +=  number_format
                sheet['M'+str(row)].alignment += alignmentNumber
                sheet['N'+str(row)] =  'NE'
                sheet['N'+str(row)].font +=  number_format
                sheet['N'+str(row)].alignment += alignmentNumber
                sheet['N'+str(row)].number_format = '#,#0.0'
                sheet['O'+str(row)] =  '0'
                sheet['O'+str(row)].font +=  number_format
                sheet['O'+str(row)].alignment += alignmentNumber
                sheet['O'+str(row)].number_format = '#,#0.0'
                sheet['P'+str(row)] =  '0'
                sheet['P'+str(row)].font +=  number_format
                sheet['P'+str(row)].alignment += alignmentNumber
                sheet['P'+str(row)].number_format = '#,#0.0'
                sheet['Q'+str(row)] =  '0'
                sheet['Q'+str(row)].font +=  number_format
                sheet['Q'+str(row)].alignment += alignmentNumber
                sheet['Q'+str(row)].number_format = '#,##0.00'
                sheet['R'+str(row)] =  '0'
                sheet['R'+str(row)].font +=  number_format
                sheet['R'+str(row)].alignment += alignmentNumber
                sheet['R'+str(row)].number_format = '#,##0.00'
                sheet['S'+str(row)] =  '0'
                sheet['S'+str(row)].font +=  number_format
                sheet['S'+str(row)].alignment += alignmentNumber
                sheet['S'+str(row)].number_format = '#,##0.00'
                sheet['T'+str(row)] =  '0'
                sheet['T'+str(row)].font +=  number_format
                sheet['T'+str(row)].alignment += alignmentNumber
                sheet['T'+str(row)].number_format = '#,##0.00'
                sheet['U'+str(row)] =  '0'
                sheet['U'+str(row)].font +=  number_format
                sheet['U'+str(row)].alignment += alignmentNumber
                sheet['U'+str(row)].number_format = '#,##0.00'

                idsperiodos =  []
                periodos = list(self.getPeriodo())
                idsperiodos.append(periodos[0][0])
                idsperiodos.append(periodos[1][0])
                idsperiodos.append(periodos[2][0])  
                vacacionesmAnt1 = 0
                vacacionesm1 = 0
                vacacionesm2 = 0
                vacacionesm3 = 0
                vacaciones = self.getVacacionesMT(empleado[1])                   
                for v in vacaciones:               
                    if v[2] in idsperiodos:
                        if idsperiodos.index(v[2]) == 0:
                            vacacionesm1 =  v[0] 
                        if idsperiodos.index(v[2]) == 1:
                            vacacionesm2 =  v[0]
                        if idsperiodos.index(v[2]) == 2:
                            vacacionesm3 =  v[0]
                    else:
                        vacacionesmAnt1 =  v[0]

                salariomes1 = 0
                horasmes1 = 0
                salariomes2 = 0
                horasmes2 = 0
                salariomes3 = 0 
                horasmes3 = 0           
                salarios = list(self.getpagoSalMT(empleado[1]))
                for sal in salarios:
                    if idsperiodos.index(sal[3]) == 0:
                        salariomes1 =  sal[0] - sal[2]
                        horasmes1 = sal[1]
                    if idsperiodos.index(sal[3]) == 1:
                        salariomes2 =  sal[0] - sal[2]
                        horasmes2 = sal[1]
                    if idsperiodos.index(sal[3]) == 2:
                        salariomes3 =  sal[0] - sal[2]
                        horasmes3 = sal[1]

                sheet['G'+str(row)] = horasmes1
                sheet['I'+str(row)] = horasmes2
                sheet['K'+str(row)] = horasmes3

                sheet['H'+str(row)] = Decimal(vacacionesmAnt1)+Decimal(vacacionesm1)+Decimal(salariomes1)
                sheet['J'+str(row)] = Decimal(vacacionesm2)+Decimal(salariomes2)
                sheet['L'+str(row)] = Decimal(vacacionesm3)+Decimal(salariomes3)

                sheet['M'+str(row)] = round((((Decimal(vacacionesmAnt1)+Decimal(vacacionesm1)+Decimal(salariomes1))+(Decimal(vacacionesm2)+Decimal(salariomes2))+(Decimal(vacacionesm3)+Decimal(salariomes3)))/3),2)
                
                mes1 = round((self.obtenerEvaCond(empleado[1],periodos[0][0])[0]),1)
                mes2 = round((self.obtenerEvaCond(empleado[1],periodos[1][0])[0]),1)
                mes3 = round((self.obtenerEvaCond(empleado[1],periodos[2][0])[0]),1)

                # if salariomes1 == 0 and mes1 != 'NE':
                #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} no tiene pago en el mes {periodos[0][1]} debe estar como NE')
                # if salariomes1 != 0 and mes1 == 'NE':
                #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} debe ser evaluado en el mes {periodos[0][1]} debido a que cuenta con un salario devengado')
                sheet['N'+str(row)] = mes1

                # if salariomes2 == 0 and mes2 != 'NE':
                #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} no tiene pago en el mes {periodos[1][1]} debe estar como NE')
                # if salariomes2 != 0 and mes2 == 'NE':
                #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} debe ser evaluado en el mes {periodos[1][1]} debido a que cuenta con un salario devengado')              
                sheet['O'+str(row)] = mes2
                
                # if salariomes3 == 0 and mes3 != 'NE':
                #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} no tiene pago en el mes {periodos[2][1]} debe estar como NE')
                # if salariomes3 != 0 and mes3 == 'NE':
                #     return messagebox.showwarning('Error en evaluación', f'El empleado {empleado[3]} debe ser evaluado en el mes {periodos[2][1]} debido a que cuenta con un salario devengado')
                sheet['P'+str(row)] = mes3
                


                promEva = self.calcCoeficienteEva(empleado[1])
                sheet['Q'+str(row)] = promEva
                if promEva <= 2:
                    sheet['R'+str(row)] = 0
                else:
                    sheet['R'+str(row)] = self.getDevengadoCalc("'"+empleado[1]+"'")[1]
                
                #Calculo del salario base de cada trabajador
                sheet['S'+str(row)] = f'=M{row}*R{row}'
                
                

                row += 1
                controw += 1
            sheet['K3'] = f'=SUM(S6:S{row+6})'
            sheet['M3'] = f'=I3/K3'
            sheet['M3'].number_format ='#,##0.00000'

            for i in range(6,row):
                sheet['T'+str(i)] = sheet['M3'].value
                sheet['U'+str(i)] = self.getDevengadoCalc("'"+sheet['C'+str(i)].value+"'")[0]#f'=rounddown((S{i}*T{i}),2)'

            
            wb.save(path)
            # separador = os.path.sep
            # dir_actual = os.path.dirname(os.path.abspath(__file__))
            # dir = separador.join(dir_actual.split(separador)[:-1])
            # dirfile = separador.join(path.split(separador))
            # command =  ['open', dir+separador+dirfile]
            # subprocess.run(command,shell=False)

            self.convert_xlsx_to_pdf(path,"utilidades_dist")
        else:
            messagebox.showwarning('Alerta','Debe realizar el cálculo resumen de las utilidades')
        
    def convert_xlsx_to_pdf(self,xlsx_file,nombreA=''):
        try:
            subprocess.run(["libreoffice24.2", "--headless", "--convert-to", "pdf", xlsx_file])
            separador = os.path.sep
            dir_actual = os.path.dirname(os.path.abspath(__file__))
            dir = separador.join(dir_actual.split(separador)[:-1])
            #dirfile = separador.join(xlsx_file.split(separador))
            command =  ['open', dir+separador+nombreA+'.pdf']
            subprocess.run(command,shell=False)
            print("Done!")

        except Exception as e:
            print("Error:", e)



    def limpiarNominaLoc(self):
        queryDemp = 'DELETE FROM postgres.public.pago_salario'
        self.cursorLoc.execute(queryDemp)
        self.connLoc.commit()

    def limpiarVacacionesLoc(self):
        queryDemp = 'DELETE FROM postgres.public.vacacionesp'
        self.cursorLoc.execute(queryDemp)
        self.connLoc.commit()

    def limpiarResumenCalcLoc(self):
        queryDemp = 'DELETE FROM postgres.public.resumen_calculo_utilidades'
        self.cursorLoc.execute(queryDemp)
        self.connLoc.commit()

    def actualizartreeEUtil1(self,event):
        self.actualizartreeEUtil()
    
    def actualizartreeEUtil(self):
        self.treeEUtil.delete(*self.treeEUtil.get_children())         
        queryEmpL=''
        if self.tx_empleado.get() != '' and self.cb_departamento.get() == '':
            queryEmpL="SELECT x.id,x.nombreap,x.ci,rc.mtsalario,rc.mtvacaciones,rc.horastt,rc.coeficienteeva_utilidades,rc.devengado,rc.descrip_coeficiente FROM postgres.public.empleado AS x INNER JOIN postgres.public.area AS a ON x.empleado_area_id = a.id INNER JOIN postgres.public.resumen_calculo_utilidades AS rc ON x.id = rc.resumen_empleado_id where x.nombreap like '%"+self.tx_empleado.get().upper()+"%' ORDER BY a.id ASC"
        elif self.cb_departamento.get() != '' and self.tx_empleado.get() == '':
            queryEmpL="SELECT x.id,x.nombreap,x.ci,rc.mtsalario,rc.mtvacaciones,rc.horastt,rc.coeficienteeva_utilidades,rc.devengado,rc.descrip_coeficiente FROM postgres.public.empleado AS x INNER JOIN postgres.public.area AS a ON x.empleado_area_id = a.id INNER JOIN postgres.public.resumen_calculo_utilidades AS rc ON x.id = rc.resumen_empleado_id where a.area = '"+self.cb_departamento.get()+"' ORDER BY a.id ASC"
        elif self.tx_empleado.get() != '' and self.cb_departamento.get() != '':
            queryEmpL="SELECT x.id,x.nombreap,x.ci,rc.mtsalario,rc.mtvacaciones,rc.horastt,rc.coeficienteeva_utilidades,rc.devengado,rc.descrip_coeficiente FROM postgres.public.empleado AS x INNER JOIN postgres.public.area AS a ON x.empleado_area_id = a.id INNER JOIN postgres.public.resumen_calculo_utilidades AS rc ON x.id = rc.resumen_empleado_id where x.nombreap like '%"+self.tx_empleado.get().upper()+"%' and a.area = '"+self.cb_departamento.get()+"' ORDER BY a.id ASC"
        else:
            queryEmpL='SELECT x.id,x.nombreap,x.ci,rc.mtsalario,rc.mtvacaciones,rc.horastt,rc.coeficienteeva_utilidades,rc.devengado,rc.descrip_coeficiente FROM postgres.public.empleado AS x INNER JOIN postgres.public.area AS a ON x.empleado_area_id = a.id INNER JOIN postgres.public.resumen_calculo_utilidades AS rc ON x.id = rc.resumen_empleado_id ORDER BY a.id ASC'
               
        self.cursorLoc.execute(queryEmpL)

        slistEmp = self.cursorLoc.fetchall()       
        for row in slistEmp:
            if row[8] != '':
                finded = False
                for einv in self.inv:
                    if einv[0] == row[0]:
                        finded = True
                if finded == False:
                    self.inv.append((row[0],row[8]))
                
            self.treeEUtil.insert('','end',values=("'"+row[0]+"'",row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8]))
        
        self.tx_total['text'] = 'Total de registros: '+str(len(self.treeEUtil.get_children()))
        if len(self.treeEUtil.get_children()) > 0:
            self.calculo_utili = True
            

    def cargarDpto(self):
        options=[]         
        queryP='SELECT x.* FROM postgres.public.area x order by area asc'
        self.cursorLoc.execute(queryP)
        slistArea=self.cursorLoc.fetchall()
        for row in slistArea:
            options.append(row[1])
        
        self.cb_departamento['values']=options

    def getDevengadoCalc(self, emp):
        query = "SELECT x.devengado,x.coeficienteeva_utilidades FROM postgres.public.resumen_calculo_utilidades x where x.resumen_empleado_id ="+emp
        self.cursorLoc.execute(query)
        return self.cursorLoc.fetchone()


    def getPeriodo(self):         
        queryP='SELECT p.* FROM postgres.public.utilidades_periodo_incluye x INNER JOIN postgres.public.periodo AS p ON x.upincluye_periodo_id = p.id order by p.id asc'
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchall()

    def getPeriodo1(self):         
        queryP='SELECT p.* FROM postgres.public.periodo AS p order by p.id asc'
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchall()

    def obtenerPerMes(self,mes):
        queryP="SELECT * FROM postgres.public.periodo where mes='"+str(mes)+"'"
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchall()
   
    def getVacaInciTrab(self,no_interno, periodo):
        queryNomina = "SELECT no_interno FROM ZUNpr.dbo.nomina_vac WHERE id_periodo = "+str(periodo)+" AND no_interno = "+no_interno+" AND deveng_vac != 0"
        self.cursorZun.execute(queryNomina)        
        idEmplwNom = self.cursorZun.fetchone()
        if idEmplwNom is not None:
            queryInci = "SELECT x.id_inci,x.id_padre FROM ZUNpr.dbo.h_incidencias x \
            WHERE x.no_interno = '"+no_interno+"' AND x.id_ppago ="+str(periodo)+" AND x.tipo = 4 ORDER BY x.id_inci DESC"
            self.cursorZun.execute(queryInci)
            result = self.cursorZun.fetchone()
            if result is not None:
                if result['id_padre']==0:
                    return result    

    def getDepartamento(self,idemp):         
        queryP="SELECT a.area  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id where emp.id = "+str(idemp)
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchone()[0]

    def getListOP(self,emp):
        queryOP= "SELECT o.monto FROM postgres.public.opago AS o WHERE opago_empleado_id="+str(emp)
        self.cursorLoc.execute(queryOP)
        return self.cursorLoc.fetchall()


    def showResumen(self):                 
        queryP="SELECT emp.id,emp.nombreap,emp.ci,a.area  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id ORDER BY a.id"
        self.cursorLoc.execute(queryP)
        sumasaldevt = 0
        empList = self.cursorLoc.fetchall()
        registro_salcalc = []
        utilidades_distrib = self.tx_distribuir.get()
        if utilidades_distrib == '':
            return messagebox.showinfo('Campo requerido','Debe indicar el monto a distribuir')
        self.limpiarResumenCalcLoc()
        sumasalcalc = 0
        coeficiente_distribuir = 0
        for emp in empList:
            mtvacaciones = 0
            mtsalario = 0
            horast = 0
            coeficienteEva = 0
            listSal = self.getpagoSalMT(emp[0])
            listVaca = self.getVacacionesMT(emp[0])
            for rowsal in listSal:
                mtsalario += rowsal[0] - rowsal[2]
                horast += rowsal[1]
            for rowvaca in listVaca:
                mtvacaciones += rowvaca[0]

            lop=self.getListOP("'"+emp[0]+"'")
            if len(lop)>0:
                for op in lop:
                   mtsalario+=op[0] 


            promEva = self.calcCoeficienteEva(emp[0])
            if promEva <= 2:
                coeficienteEva = 0
            else:
                coeficienteEva = promEva

            finded = False            
            for einv in self.inv:
                if einv[0] == emp[0]:
                    finded = True
            if finded:
                coeficienteEva = 0.0
                

            salcalc = ((Decimal(mtsalario) + Decimal(mtvacaciones))/3)*Decimal(coeficienteEva)
            registro_salcalc.append((emp[0],salcalc,mtsalario,mtvacaciones,horast))

        for empsalp in registro_salcalc:
            sumasalcalc += empsalp[1]

        coeficiente_distribuir = Decimal(utilidades_distrib)/Decimal(sumasalcalc)

        queryUpdateUD = "UPDATE postgres.public.utilidades_distribucion set monto_distribuir ="+str(self.tx_distribuir.get())+" WHERE id ="+str(self.getUtiliDist()[0])
        self.cursorLoc.execute(queryUpdateUD)
        self.connLoc.commit()
        for listE in registro_salcalc:
            saldevT = self.truncar_2_decimales((listE[1] * coeficiente_distribuir))
            sumasaldevt += saldevT
            finded = False  
            descripinv = ''          
            for einv in self.inv:
                if einv[0] == listE[0]:
                    finded = True
                    descripinv = str(einv[1])
            if finded:
                queryInsertRe = "INSERT INTO postgres.public.resumen_calculo_utilidades\
                    (resumen_empleado_id,resumen_utilidadesd_id,mtvacaciones,mtsalario,horastt,coeficienteeva_utilidades,descrip_coeficiente,devengado)\
                        VALUES ('"+str(listE[0])+"',"+str(self.getUtiliDist()[0])+","+str(round(listE[3],2))+","+str(round(listE[2],2))+","+str(round(listE[4],2))+",0.0,'"+descripinv.strip()+"',"+str(saldevT)+")"
            else:
                queryInsertRe = "INSERT INTO postgres.public.resumen_calculo_utilidades\
                    (resumen_empleado_id,resumen_utilidadesd_id,mtvacaciones,mtsalario,horastt,coeficienteeva_utilidades,descrip_coeficiente,devengado)\
                        VALUES ('"+str(listE[0])+"',"+str(self.getUtiliDist()[0])+","+str(round(listE[3],2))+","+str(round(listE[2],2))+","+str(round(listE[4],2))+","+str(self.calcCoeficienteEva(listE[0]))+",'',"+str(saldevT)+")"
           

            self.cursorLoc.execute(queryInsertRe)
            self.connLoc.commit()
        self.calculo_utili = True
        diferencia = Decimal(utilidades_distrib)-Decimal(sumasaldevt)
        
        self.tx_diferencia['text'] = 'Diferencia: '+str(round(diferencia,2))
        self.actualizartreeEUtil()


    #Redondear por defecto
    def truncar_2_decimales(self,numero):
        factor = 100
        return int(numero * factor) / factor    

    #Obtener informacion del Periodo de utilidades definido        
    def getUtiliDist(self):
        queryUD="SELECT x.* FROM postgres.public.utilidades_distribucion x"
        self.cursorLoc.execute(queryUD)
        return self.cursorLoc.fetchone()        

    def calcCoeficienteEva(self, emp):
        listPer = self.getPeriodo()
        sumeva = 0
        coeficiente = 0
        countDiv = len(listPer)
        for periodo in listPer:
            queryEva = "SELECT te.peso FROM postgres.public.tipo_evaluacion te \
            INNER JOIN postgres.public.evaluacion AS e ON te.id = e.evaluacion_tipoevaluacion_id WHERE e.evaluacion_perio_id="\
            +str(periodo[0])+" AND e.evaluacion_empleado_id = '"+emp+"'"
            self.cursorLoc.execute(queryEva)
            eva = self.cursorLoc.fetchone()
            if eva is not None:
                if eva[0] == 0:
                    countDiv -= 1
                else:
                    sumeva += eva[0]
            else:
                return messagebox.showinfo('Notificación',"'Trabajador '"+emp+"'No tiene registro de evaluación en el mes de "+str(periodo[1]))
        if countDiv > 0: 
            coeficiente = sumeva/countDiv
            if coeficiente >2:
                return coeficiente        
        return coeficiente

    def resumendetallemin(self):
        path = "file/resumen_detalle_min.xlsx"
        row = 3 
        controw = 1
        self.limpiarExcel(row,path)   
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
             
        queryP="SELECT a.area,emp.id,emp.ci,emp.nombreap,emp.escalas,emp.thoraria  FROM postgres.public.empleado emp INNER JOIN postgres.public.area AS a ON emp.empleado_area_id  = a.id ORDER BY a.id"
        self.cursorLoc.execute(queryP)
        listEmp = self.cursorLoc.fetchall()
        for empleado in listEmp:
            sheet['B'+str(row)] =  controw
            sheet['C'+str(row)] =  empleado[1]
            sheet['D'+str(row)] =  empleado[2]
            sheet['E'+str(row)] =  empleado[3]
            sheet['F'+str(row)] =  empleado[4]
            sheet['G'+str(row)] =  empleado[5]
            if self.getDestajo(empleado[1])[1] is True:
                sheet['H'+str(row)] =  'Si'
            else:
                sheet['H'+str(row)] =  'No'
                
            sheet['I'+str(row)] =  '0'
            sheet['J'+str(row)] =  '0'
            sheet['K'+str(row)] =  '0'
            sheet['L'+str(row)] =  '0'
            sheet['S'+str(row)] =  '0'
            sheet['T'+str(row)] =  '0'
            sheet['AA'+str(row)] =  '0'
            sheet['AB'+str(row)] =  '0'
            idsperiodos =  []
            periodos = list(self.getPeriodo())
            idsperiodos.append(periodos[0][0])
            idsperiodos.append(periodos[1][0])
            idsperiodos.append(periodos[2][0])  
            vacaciones = self.getVacacionesMT(empleado[1])                   
            for v in vacaciones:               
                if v[2] in idsperiodos:
                    if idsperiodos.index(v[2]) == 0:
                        sheet['K'+str(row)] =  v[1]
                        sheet['L'+str(row)] =  v[0]
                    if idsperiodos.index(v[2]) == 1:
                        sheet['S'+str(row)] =  v[1]
                        sheet['T'+str(row)] =  v[0]
                    if idsperiodos.index(v[2]) == 2:
                        sheet['AA'+str(row)] =  v[1]
                        sheet['AB'+str(row)] =  v[0]
                else:
                    sheet['I'+str(row)] =  v[1]
                    sheet['J'+str(row)] =  v[0]

            sheet['M'+str(row)] =  '0'
            sheet['N'+str(row)] =  '0'
            sheet['U'+str(row)] =  '0'
            sheet['V'+str(row)] =  '0'
            sheet['AC'+str(row)] =  '0'
            sheet['AD'+str(row)] =  '0'
            salarios = list(self.getpagoSalMT(empleado[1]))
            for sal in salarios:
                if idsperiodos.index(sal[3]) == 0:
                    sheet['M'+str(row)] =  sal[2]
                    sheet['N'+str(row)] =  sal[0]
                if idsperiodos.index(sal[3]) == 1:
                    sheet['U'+str(row)] =  sal[2]
                    sheet['V'+str(row)] =  sal[0]
                if idsperiodos.index(sal[3]) == 2:
                    sheet['AC'+str(row)] =  sal[2]
                    sheet['AD'+str(row)] =  sal[0]


            row += 1
            controw += 1
        wb.save(path)
        self.convert_xlsx_to_pdf(path,'resumen_detalle_min')
        # separador = os.path.sep
        # dir_actual = os.path.dirname(os.path.abspath(__file__))
        # dir = separador.join(dir_actual.split(separador)[:-1])
        # dirfile = separador.join(path.split(separador))
        
        # command =  ['open', dir+separador+dirfile]
        # subprocess.run(command,shell=False)



    def limpiarExcel(self,fila,url):         
        path = url
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        sheet.delete_rows(fila, sheet.max_row-1)        
        wb.save(path)




    def getpagoSalMT(self, empleado):
        mtsalario = []
        querygetSal="SELECT ps.sal_devengado,ps.destajo,ps.horast,ps.psalario_periodo_id FROM postgres.public.utilidades_periodo_incluye up \
        INNER JOIN postgres.public.pago_salario AS ps ON up.upincluye_periodo_id = ps.psalario_periodo_id where ps.psalario_empleado_id  = '"+empleado+"' order by up.upincluye_periodo_id asc" 
        self.cursorLoc.execute(querygetSal)
        salarioList = self.cursorLoc.fetchall()
        for sal in salarioList:
            mtsalario.append(((sal[0]),sal[2],sal[1],sal[3]))
        return mtsalario


    def getVacacionesMT(self, empleado):
        mtvacaciones = []
        querygetVaca="SELECT v.importe_total,v.tiempo_tota,p.id  FROM postgres.public.periodo p \
        INNER JOIN postgres.public.vacacionesp AS v ON p.id = v.vacacionesp_periodo_id where v.vacacionesp_empleado_id  = '"+empleado+"' order by p.id asc" 
        self.cursorLoc.execute(querygetVaca)
        vacacionesList = self.cursorLoc.fetchall()
        for vac in vacacionesList:
            if self.getDestajo(empleado)[1]:
                tiempo=float(vac[1])
                if self.getTarifaZun(empleado[0],vac[2]) is not None:
                    tarifaH = self.getTarifaZun(empleado[0],vac[2])[0]
                else:
                    tarifaH=float(self.getDestajo(empleado)[0])
                calcVaca = tiempo*tarifaH*8
                mtvacaciones.append((round(calcVaca,2), vac[1],vac[2]))
            else:
                mtvacaciones.append((vac[0], vac[1],vac[2]))
        return mtvacaciones

        

        

    def getDestajo(self,empleado):
        querygetSal="SELECT e.thoraria,e.destajo FROM postgres.public.empleado AS e \
            where e.id  = '"+empleado+"'"
        self.cursorLoc.execute(querygetSal)
        result = self.cursorLoc.fetchone()
        return result

    def getTarifaZun(self,empleado,periodo):
        querygetTarifa="SELECT x.tarifa_horaria FROM ZUNpr.dbo.h_empleado x \
            where x.no_interno  = '"+empleado+"' AND x.id_peri = "+str(periodo)
        self.cursorZun.execute(querygetTarifa)
        result = self.cursorZun.fetchone()
        return result

    def obtenerEvaCond(self,emp,periodo):
        queryP="SELECT te.peso FROM postgres.public.evaluacion AS e  INNER JOIN postgres.public.tipo_evaluacion AS te ON e.evaluacion_tipoevaluacion_id = te.id where e.evaluacion_empleado_id='"+str(emp)+"' and e.evaluacion_perio_id='"+str(periodo)+"'"
        #print(queryP)
        self.cursorLoc.execute(queryP)
        return self.cursorLoc.fetchone()



    

    
